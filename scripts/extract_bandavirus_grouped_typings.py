from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


DEFAULT_GROUPED_CSV = Path("database/virus/bandavirus/bandavirus_db_grouped.csv")
DEFAULT_GROUPED_FASTA = Path("database/virus/bandavirus/bandavirus_db_grouped.fasta")
DEFAULT_AF_XLSX = Path("database/virus/bandavirus/A_Fgroup/SFTSV_tree_sample_groups.xlsx")
DEFAULT_CJ_XLSX = Path("database/virus/bandavirus/CJ_group/CJ_group.xlsx")
DEFAULT_OUT_ROOT = Path("database/virus/bandavirus/grouped_typing_refs")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract A_Fgroup and CJ_group typing references from bandavirus_db_grouped resources."
    )
    parser.add_argument("--grouped-csv", type=Path, default=DEFAULT_GROUPED_CSV)
    parser.add_argument("--grouped-fasta", type=Path, default=DEFAULT_GROUPED_FASTA)
    parser.add_argument("--af-xlsx", type=Path, default=DEFAULT_AF_XLSX)
    parser.add_argument("--cj-xlsx", type=Path, default=DEFAULT_CJ_XLSX)
    parser.add_argument("--out-root", type=Path, default=DEFAULT_OUT_ROOT)
    return parser.parse_args()


def normalize_accession(value: object) -> str:
    return str(value or "").strip().split("/", 1)[0].split(".", 1)[0]


def normalize_key(text: str) -> str:
    return "".join(ch for ch in str(text).lower() if ch.isalnum())


def safe_name(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in {"_", "-", "."} else "_" for ch in value).strip("_") or "unknown"


def wrap_sequence(sequence: str, width: int = 80) -> str:
    return "\n".join(sequence[index:index + width] for index in range(0, len(sequence), width))


def normalize_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return ""
    return str(value).strip()


def read_grouped_fasta_map(path: Path) -> dict[str, dict[str, str]]:
    fasta_map: dict[str, dict[str, str]] = {}
    header = ""
    seq_lines: list[str] = []
    current_accession = ""
    current_segment = ""
    current_sample = ""
    for raw_line in path.open(encoding="utf-8", errors="ignore"):
        line = raw_line.rstrip("\n")
        if line.startswith(">"):
            if header:
                fasta_map[f"{current_accession}|{current_segment}"] = {
                    "header": header,
                    "sequence": "".join(seq_lines).upper(),
                    "accession": current_accession,
                    "segment": current_segment,
                    "sample": current_sample,
                }
            header = line[1:].strip()
            parts = [part.strip() for part in header.split("|")]
            meta: dict[str, str] = {}
            for part in parts:
                if "=" in part:
                    key, value = part.split("=", 1)
                    meta[key.strip()] = value.strip()
            current_accession = meta.get("accession", "")
            current_segment = meta.get("segment", "")
            current_sample = meta.get("sample", "")
            seq_lines = []
        else:
            seq_lines.append(line.strip())
    if header:
        fasta_map[f"{current_accession}|{current_segment}"] = {
            "header": header,
            "sequence": "".join(seq_lines).upper(),
            "accession": current_accession,
            "segment": current_segment,
            "sample": current_sample,
        }
    return fasta_map


def load_grouped_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def build_grouped_index(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    return {
        f"{(row.get('Accession') or '').strip()}|{(row.get('Assigned_Segment') or '').strip()}": row
        for row in rows
    }


def load_af_rows(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["All_segments"]
    rows: list[dict[str, str]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not any(raw):
            continue
        segment, sample_id, group = raw[:3]
        rows.append(
            {
                "segment": str(segment or "").strip(),
                "sample_id": str(sample_id or "").strip(),
                "group": str(group or "").strip(),
                "requested_accession": normalize_accession(sample_id),
            }
        )
    return rows


def resolve_af_accession(row: dict[str, str], grouped_rows: list[dict[str, str]]) -> tuple[str | None, str]:
    requested = row["requested_accession"]
    segment = row["segment"]
    for candidate in grouped_rows:
        if (candidate.get("Assigned_Segment") or "").strip() != segment:
            continue
        if (candidate.get("Accession") or "").strip() == requested:
            return requested, "direct"

    manual_map = {
        ("L", "NC018136"): "NC_018136",
        ("M", "PQ761121"): "PP761121",
        ("M", "PQ761122"): "PP761122",
        ("M", "KC473534"): "KC473541",
        ("M", "KF356640"): "KF356540",
        ("M", "OQ670930"): "JQ670930",
        ("S", "AB898541"): "AB985541",
        ("L", "SFTSZJK1"): "PX226566",
        ("M", "SFTSZJK1"): "PX226564",
        ("S", "SFTSZJK1"): "PX226562",
        ("L", "SFTSZJK2"): "PX226567",
        ("M", "SFTSZJK2"): "PX226565",
        ("S", "SFTSZJK2"): "PX226563",
    }
    mapped = manual_map.get((segment, requested))
    if mapped:
        return mapped, "manual_map"

    parts = row["sample_id"].split("/")
    label = parts[1] if len(parts) > 1 else parts[0]
    label_norm = normalize_key(label)
    hits: list[str] = []
    for candidate in grouped_rows:
        if (candidate.get("Assigned_Segment") or "").strip() != segment:
            continue
        haystacks = [
            candidate.get("Accession", ""),
            candidate.get("Isolate", ""),
            candidate.get("GenBank_Title", ""),
            candidate.get("Sample_Key", ""),
            candidate.get("Original_Header", ""),
        ]
        if label_norm and label_norm in normalize_key(" ".join(haystacks)):
            hits.append((candidate.get("Accession") or "").strip())
    hits = sorted(set(hits))
    if len(hits) == 1:
        return hits[0], "label_match"
    return None, "unresolved"


def build_af_header(info: dict[str, str], grouped_row: dict[str, str], sequence: str) -> str:
    return " | ".join(
        [
            f"scheme=A_Fgroup",
            f"segment={info['segment']}",
            f"group={info['group']}",
            f"accession={info['resolved_accession']}",
            f"sample_id={info['sample_id']}",
            f"length={len(sequence)}",
            f"isolate={(grouped_row.get('Isolate') or '').strip()}",
        ]
    )


def write_af_outputs(
    out_dir: Path,
    af_rows: list[dict[str, str]],
    grouped_rows: list[dict[str, str]],
    grouped_index: dict[str, dict[str, str]],
    fasta_map: dict[str, dict[str, str]],
) -> dict[str, int]:
    out_dir.mkdir(parents=True, exist_ok=True)
    segment_group_dir = out_dir / "segment_groups"
    segment_group_dir.mkdir(parents=True, exist_ok=True)

    combined_entries: list[str] = []
    manifest_rows: list[dict[str, str]] = []
    missing_rows: list[dict[str, str]] = []
    grouped_entries: dict[tuple[str, str], list[str]] = defaultdict(list)
    grouped_sample_ids: dict[tuple[str, str], list[str]] = defaultdict(list)

    for row in af_rows:
        resolved_accession, resolve_mode = resolve_af_accession(row, grouped_rows)
        key = f"{resolved_accession}|{row['segment']}" if resolved_accession else ""
        if not resolved_accession or key not in grouped_index or key not in fasta_map:
            missing_rows.append(
                {
                    "segment": row["segment"],
                    "group": row["group"],
                    "sample_id": row["sample_id"],
                    "requested_accession": row["requested_accession"],
                    "resolve_mode": resolve_mode,
                }
            )
            continue
        grouped_row = grouped_index[key]
        sequence = fasta_map[key]["sequence"]
        info = dict(row)
        info["resolved_accession"] = resolved_accession
        header = build_af_header(info, grouped_row, sequence)
        fasta_entry = f">{header}\n{wrap_sequence(sequence)}"
        combined_entries.append(fasta_entry)
        grouped_entries[(row["segment"], row["group"])].append(fasta_entry)
        grouped_sample_ids[(row["segment"], row["group"])].append(row["sample_id"])

        manifest_row = dict(grouped_row)
        manifest_row.update(
            {
                "Tree_Segment": row["segment"],
                "Tree_Group": row["group"],
                "Tree_Sample_ID": row["sample_id"],
                "Tree_Requested_Accession": row["requested_accession"],
                "Tree_Resolved_Accession": resolved_accession,
                "Resolve_Mode": resolve_mode,
                "Selected_From": "bandavirus_db_grouped.fasta",
                "Selected_Header": header,
                "Sequence_Length_From_FASTA": str(len(sequence)),
            }
        )
        manifest_rows.append(manifest_row)

    (out_dir / "SFTSV_tree_group_segments.fasta").write_text("\n".join(combined_entries).strip() + "\n", encoding="utf-8")

    manifest_columns = list(grouped_rows[0].keys()) + [
        "Tree_Segment",
        "Tree_Group",
        "Tree_Sample_ID",
        "Tree_Requested_Accession",
        "Tree_Resolved_Accession",
        "Resolve_Mode",
        "Selected_From",
        "Selected_Header",
        "Sequence_Length_From_FASTA",
    ]
    with (out_dir / "SFTSV_tree_group_segments_manifest.tsv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=manifest_columns, delimiter="\t")
        writer.writeheader()
        writer.writerows(manifest_rows)

    summary_rows: list[dict[str, str]] = []
    for (segment, group), entries in sorted(grouped_entries.items()):
        fasta_path = segment_group_dir / f"{safe_name(segment)}_{safe_name(group)}.fasta"
        fasta_path.write_text("\n".join(entries).strip() + "\n", encoding="utf-8")
        summary_rows.append(
            {
                "segment": segment,
                "group": group,
                "record_count": str(len(entries)),
                "sample_ids": " ; ".join(grouped_sample_ids[(segment, group)]),
                "fasta_path": str(fasta_path.resolve()),
            }
        )

    with (out_dir / "SFTSV_tree_group_summary.tsv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["segment", "group", "record_count", "sample_ids", "fasta_path"], delimiter="\t")
        writer.writeheader()
        writer.writerows(summary_rows)

    with (out_dir / "SFTSV_tree_group_missing.tsv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["segment", "group", "sample_id", "requested_accession", "resolve_mode"], delimiter="\t")
        writer.writeheader()
        writer.writerows(missing_rows)

    return {"records": len(af_rows), "resolved": len(manifest_rows), "missing": len(missing_rows), "groups": len(grouped_entries)}


def load_cj_rows(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict[str, str]] = []
    for raw_row in ws.iter_rows(min_row=4, values_only=True):
        if not any(raw_row):
            continue
        viral_strain, isolation_source, isolation_date, geographic_origin, l_acc, l_geno, m_acc, m_geno, s_acc, s_geno = raw_row[:10]
        common = {
            "viral_strain": str(viral_strain or "").strip(),
            "isolation_source": str(isolation_source or "").strip(),
            "isolation_date": normalize_date(isolation_date),
            "geographic_origin": str(geographic_origin or "").strip(),
        }
        for segment, accession, genotype in (("L", l_acc, l_geno), ("M", m_acc, m_geno), ("S", s_acc, s_geno)):
            rows.append(
                {
                    **common,
                    "segment": segment,
                    "requested_accession": normalize_accession(accession),
                    "genotype": str(genotype or "").strip(),
                }
            )
    return rows


def build_cj_header(info: dict[str, str], grouped_row: dict[str, str], sequence: str) -> str:
    return " | ".join(
        [
            "scheme=CJ_group",
            f"strain={info['viral_strain']}",
            f"segment={info['segment']}",
            f"genotype={info['genotype']}",
            f"accession={info['requested_accession']}",
            f"length={len(sequence)}",
            f"isolate={(grouped_row.get('Isolate') or '').strip()}",
        ]
    )


def write_cj_outputs(
    out_dir: Path,
    cj_rows: list[dict[str, str]],
    grouped_rows: list[dict[str, str]],
    grouped_index: dict[str, dict[str, str]],
    fasta_map: dict[str, dict[str, str]],
) -> dict[str, int]:
    out_dir.mkdir(parents=True, exist_ok=True)
    subtype_dir = out_dir / "subtypes"
    strain_dir = out_dir / "strains"
    subtype_dir.mkdir(parents=True, exist_ok=True)
    strain_dir.mkdir(parents=True, exist_ok=True)

    combined_entries: list[str] = []
    manifest_rows: list[dict[str, str]] = []
    grouped_entries: dict[tuple[str, str], list[str]] = defaultdict(list)
    grouped_accessions: dict[tuple[str, str], list[str]] = defaultdict(list)
    strain_entries: dict[str, list[str]] = defaultdict(list)
    missing_rows: list[dict[str, str]] = []

    for row in cj_rows:
        key = f"{row['requested_accession']}|{row['segment']}"
        if key not in grouped_index or key not in fasta_map:
            missing_rows.append(row)
            continue
        grouped_row = grouped_index[key]
        sequence = fasta_map[key]["sequence"]
        header = build_cj_header(row, grouped_row, sequence)
        fasta_entry = f">{header}\n{wrap_sequence(sequence)}"
        combined_entries.append(fasta_entry)
        grouped_entries[(row["segment"], row["genotype"])].append(fasta_entry)
        grouped_accessions[(row["segment"], row["genotype"])].append(row["requested_accession"])
        strain_entries[row["viral_strain"]].append(fasta_entry)

        manifest_row = dict(grouped_row)
        manifest_row.update(
            {
                "CJ_Viral_Strain": row["viral_strain"],
                "CJ_Isolation_Source": row["isolation_source"],
                "CJ_Isolation_Date": row["isolation_date"],
                "CJ_Geographic_Origin": row["geographic_origin"],
                "CJ_Segment": row["segment"],
                "CJ_Genotype": row["genotype"],
                "Selected_From": "bandavirus_db_grouped.fasta",
                "Selected_Header": header,
                "Sequence_Length_From_FASTA": str(len(sequence)),
            }
        )
        manifest_rows.append(manifest_row)

    (out_dir / "CJ_group_segment_genomes.fasta").write_text("\n".join(combined_entries).strip() + "\n", encoding="utf-8")

    manifest_columns = list(grouped_rows[0].keys()) + [
        "CJ_Viral_Strain",
        "CJ_Isolation_Source",
        "CJ_Isolation_Date",
        "CJ_Geographic_Origin",
        "CJ_Segment",
        "CJ_Genotype",
        "Selected_From",
        "Selected_Header",
        "Sequence_Length_From_FASTA",
    ]
    with (out_dir / "CJ_group_segment_genomes_manifest.tsv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=manifest_columns, delimiter="\t")
        writer.writeheader()
        writer.writerows(manifest_rows)

    subtype_summary_rows: list[dict[str, str]] = []
    for (segment, genotype), entries in sorted(grouped_entries.items()):
        fasta_path = subtype_dir / f"{safe_name(segment)}_{safe_name(genotype)}.fasta"
        fasta_path.write_text("\n".join(entries).strip() + "\n", encoding="utf-8")
        subtype_summary_rows.append(
            {
                "segment": segment,
                "genotype": genotype,
                "record_count": str(len(entries)),
                "accessions": ",".join(grouped_accessions[(segment, genotype)]),
                "fasta_path": str(fasta_path.resolve()),
            }
        )

    with (out_dir / "CJ_group_subtype_manifest.tsv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["segment", "genotype", "record_count", "accessions", "fasta_path"], delimiter="\t")
        writer.writeheader()
        writer.writerows(subtype_summary_rows)

    for strain, entries in strain_entries.items():
        (strain_dir / f"{safe_name(strain)}.fasta").write_text("\n".join(entries).strip() + "\n", encoding="utf-8")

    with (out_dir / "CJ_group_missing.tsv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "viral_strain",
                "isolation_source",
                "isolation_date",
                "geographic_origin",
                "segment",
                "requested_accession",
                "genotype",
            ],
            delimiter="\t",
        )
        writer.writeheader()
        writer.writerows(missing_rows)

    return {"records": len(cj_rows), "resolved": len(manifest_rows), "missing": len(missing_rows), "groups": len(grouped_entries)}


def main() -> int:
    args = parse_args()
    grouped_csv = args.grouped_csv.expanduser().resolve()
    grouped_fasta = args.grouped_fasta.expanduser().resolve()
    af_xlsx = args.af_xlsx.expanduser().resolve()
    cj_xlsx = args.cj_xlsx.expanduser().resolve()
    out_root = args.out_root.expanduser().resolve()
    out_root.mkdir(parents=True, exist_ok=True)

    grouped_rows = load_grouped_rows(grouped_csv)
    grouped_index = build_grouped_index(grouped_rows)
    fasta_map = read_grouped_fasta_map(grouped_fasta)

    af_stats = write_af_outputs(out_root / "A_Fgroup", load_af_rows(af_xlsx), grouped_rows, grouped_index, fasta_map)
    cj_stats = write_cj_outputs(out_root / "CJ_group", load_cj_rows(cj_xlsx), grouped_rows, grouped_index, fasta_map)

    print(f"A_Fgroup_records\t{af_stats['records']}")
    print(f"A_Fgroup_resolved\t{af_stats['resolved']}")
    print(f"A_Fgroup_missing\t{af_stats['missing']}")
    print(f"A_Fgroup_groups\t{af_stats['groups']}")
    print(f"CJ_group_records\t{cj_stats['records']}")
    print(f"CJ_group_resolved\t{cj_stats['resolved']}")
    print(f"CJ_group_missing\t{cj_stats['missing']}")
    print(f"CJ_group_groups\t{cj_stats['groups']}")
    print(f"out_root\t{out_root}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
