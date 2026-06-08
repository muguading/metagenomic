#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import http.client
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX_PATH = ROOT / "database" / "virus" / "Hepatovirus" / "typing.xlsx"
DEFAULT_OUT_DIR = ROOT / "database" / "virus" / "Hepatovirus" / "reference_genomes"

EUTILS_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
SVIEWER_URL = "https://www.ncbi.nlm.nih.gov/sviewer/viewer.fcgi"
USER_AGENT = "metagenomic-hepatovirus-reference-downloader/1.0"
REQUEST_SLEEP_SECONDS = 0.34


def fetch_text(url: str, params: dict[str, str], retries: int = 6, timeout: int = 90) -> str:
    query = urllib.parse.urlencode(params)
    full_url = f"{url}?{query}"
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        request = urllib.request.Request(full_url, headers={"User-Agent": USER_AGENT})
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                return response.read().decode("utf-8", errors="replace")
        except (
            urllib.error.HTTPError,
            urllib.error.URLError,
            TimeoutError,
            http.client.IncompleteRead,
            ConnectionError,
        ) as exc:
            last_error = exc
            if attempt == retries:
                break
            delay = max(1.5, attempt * 2.0)
            if isinstance(exc, urllib.error.HTTPError) and exc.code == 429:
                delay = max(delay, 6.0 * attempt)
            print(f"[retry] {full_url} attempt {attempt}/{retries}: {exc}", file=sys.stderr)
            time.sleep(delay)
    raise RuntimeError(f"failed to fetch {full_url}: {last_error}") from last_error


def read_typing_rows(xlsx_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        if not values or not values[4]:
            continue
        rows.append(
            {
                "genus": str(values[0] or "").strip(),
                "species": str(values[1] or "").strip(),
                "virus_name": str(values[2] or "").strip(),
                "isolate": str(values[3] or "").strip(),
                "accession": str(values[4] or "").strip(),
                "available_sequence": str(values[5] or "").strip(),
                "abbrev": str(values[6] or "").strip(),
            }
        )
    return rows


def sequence_length_from_fasta(fasta_path: Path) -> tuple[str, str]:
    header = ""
    sequence_length = 0
    with fasta_path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line in handle:
            if line.startswith(">") and not header:
                header = line[1:].strip()
            elif line and not line.startswith(">"):
                sequence_length += len(line.strip())
    return header, str(sequence_length)


def download_one(row: dict[str, str], out_dir: Path) -> dict[str, str]:
    accession = row["accession"]
    fasta_path = out_dir / f"{accession}.fasta"
    gff_path = out_dir / "gff3" / f"{accession}.gff3"
    fasta_path.parent.mkdir(parents=True, exist_ok=True)
    gff_path.parent.mkdir(parents=True, exist_ok=True)

    fasta_status = "cached"
    gff_status = "cached"

    if not fasta_path.is_file() or fasta_path.stat().st_size == 0:
        fasta_text = fetch_text(
            EUTILS_URL,
            {
                "db": "nuccore",
                "id": accession,
                "rettype": "fasta",
                "retmode": "text",
            },
        )
        if not fasta_text.lstrip().startswith(">"):
            raise RuntimeError(f"{accession}: invalid FASTA response")
        fasta_path.write_text(fasta_text.rstrip() + "\n", encoding="utf-8")
        fasta_status = "downloaded"
        time.sleep(REQUEST_SLEEP_SECONDS)

    if not gff_path.is_file() or gff_path.stat().st_size == 0:
        gff_text = fetch_text(
            SVIEWER_URL,
            {
                "id": accession,
                "db": "nuccore",
                "report": "gff3",
                "retmode": "text",
            },
        )
        if "##gff-version" not in gff_text:
            raise RuntimeError(f"{accession}: invalid GFF3 response")
        gff_path.write_text(gff_text.rstrip() + "\n", encoding="utf-8")
        gff_status = "downloaded"
        time.sleep(REQUEST_SLEEP_SECONDS)

    header, sequence_length = sequence_length_from_fasta(fasta_path)
    result = dict(row)
    result.update(
        {
            "header": header,
            "sequence_length": sequence_length,
            "fasta_path": str(fasta_path),
            "gff3_path": str(gff_path),
            "fasta_status": fasta_status,
            "gff3_status": gff_status,
            "status": "ok",
        }
    )
    return result


def write_outputs(rows: list[dict[str, str]], out_dir: Path) -> None:
    combined_fasta_path = out_dir / "hepatovirus_typing_reference_genomes.fasta"
    manifest_path = out_dir / "hepatovirus_typing_reference_genomes_manifest.tsv"

    with combined_fasta_path.open("w", encoding="utf-8") as output:
        for row in rows:
            fasta_path = Path(row["fasta_path"])
            output.write(fasta_path.read_text(encoding="utf-8", errors="ignore").rstrip() + "\n")

    fieldnames = [
        "genus",
        "species",
        "virus_name",
        "isolate",
        "accession",
        "available_sequence",
        "abbrev",
        "header",
        "sequence_length",
        "fasta_path",
        "gff3_path",
        "fasta_status",
        "gff3_status",
        "status",
    ]
    with manifest_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def main() -> int:
    parser = argparse.ArgumentParser(description="Download Hepatovirus typing reference genomes from NCBI")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX_PATH, help="typing.xlsx path")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR, help="Output directory")
    parser.add_argument("--workers", type=int, default=3, help="Concurrent download workers")
    args = parser.parse_args()

    xlsx_path = args.xlsx.expanduser().resolve()
    out_dir = args.out_dir.expanduser().resolve()
    rows = read_typing_rows(xlsx_path)
    if not rows:
        print(f"No accessions found in {xlsx_path}", file=sys.stderr)
        return 1

    results: list[dict[str, str]] = []
    errors: list[tuple[str, str]] = []
    max_workers = max(1, min(args.workers, 6))

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {executor.submit(download_one, row, out_dir): row["accession"] for row in rows}
        for future in as_completed(future_map):
            accession = future_map[future]
            try:
                result = future.result()
                results.append(result)
                print(f"[ok] {accession}")
            except Exception as exc:  # noqa: BLE001
                errors.append((accession, str(exc)))
                print(f"[fail] {accession}: {exc}", file=sys.stderr)

    results.sort(key=lambda item: item["accession"])
    if results:
        write_outputs(results, out_dir)

    if errors:
        print(f"{len(errors)} accessions failed", file=sys.stderr)
        for accession, message in errors:
            print(f"  - {accession}: {message}", file=sys.stderr)
        return 2

    print(f"Downloaded {len(results)} Hepatovirus references to {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
