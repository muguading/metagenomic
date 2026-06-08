#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import subprocess
from collections import defaultdict
from pathlib import Path

from Bio import SeqIO
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord
from openpyxl import load_workbook


ROOT = Path("/Users/wuhhh/Desktop/徐老师/代码/metagenomic")
TYPING_XLSX = ROOT / "database" / "virus" / "enterovirus" / "typing.xlsx"
REF_DIR = ROOT / "database" / "virus" / "enterovirus" / "reference_genomes"
VP1_MANIFEST = REF_DIR / "enterovirus_vp1_references.tsv"
SUPPLEMENT_MANIFEST = REF_DIR / "abcd_vp1" / "enterovirus_abcd_complete_genomes_expanded_manifest.tsv"
OUT_DIR = REF_DIR / "abcd_vp1"
OUT_FASTA = OUT_DIR / "enterovirus_abcd_vp1.fasta"
OUT_TSV = OUT_DIR / "enterovirus_abcd_vp1.tsv"
OUT_COMPLETENESS = OUT_DIR / "enterovirus_abcd_subtype_completeness.tsv"
INFERRED_DIR = OUT_DIR / "inferred_vp1"
BLASTN_BIN = ROOT / "soft" / "ncbi-blast" / "bin" / "blastn"

SPECIES_TO_GROUP = {
    "Enterovirus alphacoxsackie": "A",
    "Enterovirus alpharhino": "A",
    "Enterovirus betacoxsackie": "B",
    "Enterovirus betarhino": "B",
    "Enterovirus coxsackiepol": "C",
    "Enterovirus cerhino": "C",
    "Enterovirus deconjuncti": "D",
}


def load_typing_rows():
    wb = load_workbook(TYPING_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(x or "").strip() for x in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    result = []
    for row in rows[1:]:
        accession = str(row[idx["Accession"]] or "").strip()
        if not accession or accession.startswith("partial:"):
            continue
        species = str(row[idx["Species"]] or "").strip()
        big_group = SPECIES_TO_GROUP.get(species)
        if not big_group:
            continue
        result.append(
            {
                "accession": accession.split(".", 1)[0],
                "accession_full": accession,
                "genus": str(row[idx["Genus"]] or "").strip(),
                "species": species,
                "virus_name": str(row[idx["Virus Name"]] or "").strip(),
                "isolate": str(row[idx["Isolate"]] or "").strip(),
                "abbrev": str(row[idx["Abbrev."]] or "").strip(),
                "available": str(row[idx["Available"]] or "").strip(),
                "big_group": big_group,
            }
        )
    return result


def load_existing_vp1_rows():
    rows = []
    with VP1_MANIFEST.open("r", encoding="utf-8", errors="ignore") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            rows.append(row)
    return rows


def load_supplement_rows():
    if not SUPPLEMENT_MANIFEST.is_file():
        return []
    rows = []
    with SUPPLEMENT_MANIFEST.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        for row in csv.DictReader(handle, delimiter="\t"):
            accession = str(row.get("accession") or "").strip()
            if not accession:
                continue
            rows.append(
                {
                    "genus": str(row.get("genus") or "").strip(),
                    "species": str(row.get("species") or "").strip(),
                    "virus_name": str(row.get("virus_name") or "").strip(),
                    "isolate": str(row.get("isolate") or row.get("title") or "").strip(),
                    "accession": accession,
                    "accession_full": str(row.get("accession_full") or accession).strip(),
                    "available": str(row.get("available") or "Complete genome").strip(),
                    "abbrev": str(row.get("abbrev") or "").strip(),
                    "big_group": str(row.get("big_group") or SPECIES_TO_GROUP.get(str(row.get("species") or "").strip(), "")).strip(),
                }
            )
    return rows


def read_single_fasta(path: Path) -> SeqRecord:
    records = list(SeqIO.parse(str(path), "fasta"))
    if not records:
        raise RuntimeError(f"No FASTA records found in {path}")
    return records[0]


def build_query_fastas(existing_rows_by_group):
    query_fastas = {}
    for group, rows in existing_rows_by_group.items():
        query_path = OUT_DIR / f"group_{group}.query.fa"
        records = []
        for row in rows:
            try:
                vp1_len = int(str(row.get("vp1_length") or "0"))
            except ValueError:
                vp1_len = 0
            if vp1_len < 650 or vp1_len > 1200:
                continue
            vp1_path = Path(row["vp1_fasta_path"])
            if not vp1_path.is_file():
                continue
            record = read_single_fasta(vp1_path)
            record.id = row["accession"]
            record.name = row["accession"]
            record.description = row.get("abbrev") or row["accession"]
            records.append(record)
        SeqIO.write(records, str(query_path), "fasta")
        query_fastas[group] = query_path
    return query_fastas


def run_blast_best_hit(query_fasta: Path, subject_fasta: Path):
    if not query_fasta.is_file() or query_fasta.stat().st_size == 0:
        return None
    cmd = [
        str(BLASTN_BIN),
        "-task",
        "blastn",
        "-query",
        str(query_fasta),
        "-subject",
        str(subject_fasta),
        "-outfmt",
        "6 qseqid qlen sseqid slen length pident bitscore evalue qstart qend sstart send qcovhsp",
        "-max_hsps",
        "1",
    ]
    completed = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=True)
    best = None
    for line in completed.stdout.splitlines():
        if not line.strip():
            continue
        parts = line.rstrip("\n").split("\t")
        if len(parts) != 13:
            continue
        row = {
            "qseqid": parts[0],
            "qlen": int(float(parts[1])),
            "sseqid": parts[2],
            "slen": int(float(parts[3])),
            "length": int(float(parts[4])),
            "pident": float(parts[5]),
            "bitscore": float(parts[6]),
            "evalue": parts[7],
            "qstart": int(float(parts[8])),
            "qend": int(float(parts[9])),
            "sstart": int(float(parts[10])),
            "send": int(float(parts[11])),
            "qcovhsp": float(parts[12]),
        }
        if best is None:
            best = row
            continue
        score = (row["bitscore"], row["qcovhsp"], row["pident"], row["length"])
        best_score = (best["bitscore"], best["qcovhsp"], best["pident"], best["length"])
        if score > best_score:
            best = row
    return best


def infer_missing_rows(typing_rows, existing_rows):
    existing_by_accession = {row["accession"]: row for row in existing_rows}
    existing_rows_by_group = defaultdict(list)
    for row in existing_rows:
        if row.get("available") != "Complete genome":
            continue
        group = SPECIES_TO_GROUP.get(row.get("species") or "")
        if group:
            existing_rows_by_group[group].append(row)

    query_fastas = build_query_fastas(existing_rows_by_group)
    inferred_rows = []
    for meta in typing_rows:
        if meta["available"] != "Complete genome":
            continue
        if meta["accession"] in existing_by_accession:
            continue
        group = meta["big_group"]
        query_fasta = query_fastas.get(group)
        subject_fasta = REF_DIR / f"{meta['accession']}.fasta"
        if not query_fasta or not subject_fasta.is_file():
            continue
        best = run_blast_best_hit(query_fasta, subject_fasta)
        if best is None:
            continue
        if best["qcovhsp"] < 70 or best["pident"] < 70 or best["length"] < 650:
            continue
        record = read_single_fasta(subject_fasta)
        start = min(best["sstart"], best["send"])
        end = max(best["sstart"], best["send"])
        strand = "+" if best["sstart"] <= best["send"] else "-"
        seq = str(record.seq)[start - 1:end]
        if strand == "-":
            seq = str(Seq(seq).reverse_complement())
        inferred_path = INFERRED_DIR / f"{meta['accession']}.vp1.fasta"
        out_record = SeqRecord(
            Seq(seq),
            id=meta["accession"],
            description=f"{record.description} | inferred VP1 {start}-{end} ({strand}) | template {best['qseqid']}",
        )
        SeqIO.write([out_record], str(inferred_path), "fasta")
        inferred_rows.append(
            {
                "accession": meta["accession"],
                "accession_full": meta["accession_full"],
                "abbrev": meta["abbrev"],
                "virus_name": meta["virus_name"],
                "species": meta["species"],
                "genus": meta["genus"],
                "isolate": meta["isolate"],
                "available": meta["available"],
                "big_group": group,
                "header": record.description,
                "vp1_start": str(start),
                "vp1_end": str(end),
                "vp1_strand": strand,
                "vp1_length": str(len(seq)),
                "feature_type": "inferred_homology",
                "product": "VP1",
                "gene": "VP1",
                "fasta_path": str(subject_fasta),
                "gff_path": "",
                "vp1_fasta_path": str(inferred_path),
                "template_accession": best["qseqid"],
                "blast_pident": f"{best['pident']:.2f}",
                "blast_qcovhsp": f"{best['qcovhsp']:.2f}",
                "blast_bitscore": f"{best['bitscore']:.1f}",
                "evidence": "same_big_group_blastn",
            }
        )
    return inferred_rows


def build_outputs():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    INFERRED_DIR.mkdir(parents=True, exist_ok=True)
    typing_rows = load_typing_rows()
    supplement_rows = load_supplement_rows()
    typing_by_accession = {row["accession"]: row for row in typing_rows}
    for row in supplement_rows:
        typing_by_accession.setdefault(row["accession"], row)
    all_reference_rows = sorted(typing_by_accession.values(), key=lambda x: (x.get("big_group", ""), x.get("abbrev", ""), x.get("accession", "")))
    existing_rows = [row for row in load_existing_vp1_rows() if row.get("available") == "Complete genome"]
    inferred_rows = infer_missing_rows(all_reference_rows, existing_rows)

    combined_by_accession = {}
    for row in existing_rows:
        row = dict(row)
        row["big_group"] = SPECIES_TO_GROUP.get(row.get("species") or "", "")
        row.setdefault("template_accession", "")
        row.setdefault("blast_pident", "")
        row.setdefault("blast_qcovhsp", "")
        row.setdefault("blast_bitscore", "")
        row.setdefault("evidence", "existing_gff")
        combined_by_accession[row["accession"]] = row
    for row in inferred_rows:
        combined_by_accession[row["accession"]] = row

    combined_rows = sorted(
        [row for row in combined_by_accession.values() if row.get("big_group") in {"A", "B", "C", "D"}],
        key=lambda x: (x.get("big_group", ""), x.get("abbrev", ""), x.get("accession", "")),
    )

    merged_records = []
    for row in combined_rows:
        vp1_path = Path(row["vp1_fasta_path"])
        if vp1_path.is_file():
            merged_records.extend(list(SeqIO.parse(str(vp1_path), "fasta")))
    SeqIO.write(merged_records, str(OUT_FASTA), "fasta")

    fieldnames = [
        "accession",
        "accession_full",
        "abbrev",
        "virus_name",
        "species",
        "genus",
        "isolate",
        "available",
        "big_group",
        "header",
        "vp1_start",
        "vp1_end",
        "vp1_strand",
        "vp1_length",
        "feature_type",
        "product",
        "gene",
        "fasta_path",
        "gff_path",
        "vp1_fasta_path",
        "template_accession",
        "blast_pident",
        "blast_qcovhsp",
        "blast_bitscore",
        "evidence",
    ]
    with OUT_TSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in combined_rows:
            writer.writerow(row)

    subtype_meta = {}
    complete_accessions_by_subtype = defaultdict(list)
    vp1_accessions_by_subtype = defaultdict(list)
    for row in all_reference_rows:
        subtype_meta.setdefault(
            row["abbrev"],
            {
                "abbrev": row["abbrev"],
                "virus_name": row["virus_name"],
                "species": row["species"],
                "big_group": row["big_group"],
            },
        )
        if row["available"] == "Complete genome":
            complete_accessions_by_subtype[row["abbrev"]].append(row["accession"])
    for row in combined_rows:
        vp1_accessions_by_subtype[row["abbrev"]].append(row["accession"])

    completeness_rows = []
    for abbrev, meta in sorted(subtype_meta.items(), key=lambda item: (item[1]["big_group"], item[0])):
        complete_accessions = sorted(set(complete_accessions_by_subtype.get(abbrev, [])))
        vp1_accessions = sorted(set(vp1_accessions_by_subtype.get(abbrev, [])))
        if not complete_accessions:
            status = "no_complete_genome"
            note = "typing.xlsx 中未见该亚型完整基因组"
        elif not vp1_accessions:
            status = "complete_genome_but_no_vp1"
            note = "存在完整基因组，但当前未获得 VP1"
        else:
            status = "ready"
            note = "存在完整基因组且已获得 VP1"
        completeness_rows.append(
            {
                "big_group": meta["big_group"],
                "abbrev": abbrev,
                "virus_name": meta["virus_name"],
                "species": meta["species"],
                "complete_genome_count": len(complete_accessions),
                "complete_genome_accessions": ",".join(complete_accessions),
                "vp1_count": len(vp1_accessions),
                "vp1_accessions": ",".join(vp1_accessions),
                "status": status,
                "note": note,
            }
        )

    with OUT_COMPLETENESS.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "big_group",
                "abbrev",
                "virus_name",
                "species",
                "complete_genome_count",
                "complete_genome_accessions",
                "vp1_count",
                "vp1_accessions",
                "status",
                "note",
            ],
            delimiter="\t",
        )
        writer.writeheader()
        for row in completeness_rows:
            writer.writerow(row)

    print(f"ABCD complete-genome VP1 references: {len(combined_rows)}")
    print(f"Inferred additional VP1 references: {len(inferred_rows)}")
    print(f"Output FASTA: {OUT_FASTA}")
    print(f"Output TSV: {OUT_TSV}")
    print(f"Completeness TSV: {OUT_COMPLETENESS}")


if __name__ == "__main__":
    build_outputs()
