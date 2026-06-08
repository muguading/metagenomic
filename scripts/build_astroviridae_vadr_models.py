#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import os
import shutil
import subprocess
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/wuhhh/Desktop/徐老师/代码/metagenomic")
SOFT_DIR = ROOT / "soft"
TYPING_XLSX = ROOT / "database" / "virus" / "astroviridae" / "typing.xlsx"
REF_DIR = ROOT / "database" / "virus" / "astroviridae" / "reference_genomes"
MODEL_ROOT = SOFT_DIR / "vadr-models-astro"
WORK_ROOT = ROOT / "tmp" / "astroviridae_vadr_build"
NCOV_PERL = Path("/opt/homebrew/Caskroom/mambaforge/base/envs/ncov/bin/perl")

GENUS_CONFIG = {
    "Avastrovirus": {"model_key": "avastrovirus", "group": "Avastrovirus"},
    "Mamastrovirus": {"model_key": "mamastrovirus", "group": "Mamastrovirus"},
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build VADR model libraries for Avastrovirus and Mamastrovirus.")
    parser.add_argument(
        "--genus",
        choices=["Avastrovirus", "Mamastrovirus", "all"],
        default="all",
        help="Restrict build to one genus. Default: all",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite existing per-accession and genus-level outputs.",
    )
    parser.add_argument(
        "--sleep-seconds",
        type=float,
        default=0.34,
        help="Sleep interval between GenBank downloads. Default: 0.34",
    )
    return parser.parse_args()


def load_complete_rows() -> list[dict[str, str]]:
    workbook = load_workbook(TYPING_XLSX, read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        if not values:
            continue
        accession = str(values[4] or "").strip()
        available = str(values[5] or "").strip()
        genus = str(values[0] or "").strip()
        if not accession or available != "Complete genome" or genus not in GENUS_CONFIG:
            continue
        accession_root = accession.split(".", 1)[0]
        rows.append(
            {
                "genus": genus,
                "species": str(values[1] or "").strip(),
                "virus_name": str(values[2] or "").strip(),
                "isolate": str(values[3] or "").strip(),
                "accession": accession_root,
                "accession_full": accession,
                "available_sequence": available,
                "abbrev": str(values[6] or "").strip(),
                "fasta_path": str((REF_DIR / f"{accession_root}.fasta").resolve()),
            }
        )
    rows.sort(key=lambda row: (row["genus"], row["accession"]))
    return rows


def build_vadr_env() -> dict[str, str]:
    vadr_root = SOFT_DIR.resolve()
    vadr_scripts_dir = (vadr_root / "vadr").resolve()
    infernal_bin_dir = (vadr_root / "infernal" / "binaries").resolve()
    bio_easel_dir = (vadr_root / "Bio-Easel-ncov").resolve()
    sequip_dir = (vadr_root / "sequip").resolve()
    blast_bin_dir = (vadr_root / "ncbi-blast" / "bin").resolve()
    fasta_bin_dir = (vadr_root / "fasta" / "bin").resolve()
    minimap2_dir = (vadr_root / "minimap2").resolve()
    local_perl_lib = (WORK_ROOT / "perl_lib").resolve()
    (local_perl_lib / "Mozilla").mkdir(parents=True, exist_ok=True)
    (local_perl_lib / "LWP" / "Protocol").mkdir(parents=True, exist_ok=True)
    (local_perl_lib / "Mozilla" / "CA.pm").write_text(
        "package Mozilla::CA;\nour $VERSION = q(0.01);\n1;\n",
        encoding="utf-8",
    )
    (local_perl_lib / "LWP" / "Protocol" / "https.pm").write_text(
        "package LWP::Protocol::https;\nour $VERSION = q(6.10);\n1;\n",
        encoding="utf-8",
    )

    inherited_path = str(os.environ.get("PATH") or "").strip()
    inherited_perl5lib = str(os.environ.get("PERL5LIB") or "").strip()
    return {
        **os.environ,
        "VADRINSTALLDIR": str(vadr_root),
        "VADRSCRIPTSDIR": str(vadr_scripts_dir),
        "VADRCONFIGFILE": str((vadr_scripts_dir / "vadr.config").resolve()),
        "VADRINFERNALDIR": str(infernal_bin_dir),
        "VADREASELDIR": str(infernal_bin_dir),
        "VADRHMMERDIR": str(infernal_bin_dir),
        "VADRBIOEASELDIR": str(bio_easel_dir),
        "VADRSEQUIPDIR": str(sequip_dir),
        "VADRBLASTDIR": str(blast_bin_dir),
        "VADRFASTADIR": str(fasta_bin_dir),
        "VADRMINIMAP2DIR": str(minimap2_dir),
        "PERL5LIB": os.pathsep.join(
            [
                str(local_perl_lib),
                str(vadr_scripts_dir),
                str(sequip_dir),
                str((bio_easel_dir / "blib" / "lib").resolve()),
                str((bio_easel_dir / "blib" / "arch").resolve()),
            ]
            + ([inherited_perl5lib] if inherited_perl5lib else [])
        ),
        "PATH": os.pathsep.join(
            [
                str(NCOV_PERL.parent),
                "/usr/bin",
                "/bin",
                str(vadr_scripts_dir),
                str(blast_bin_dir),
                str(fasta_bin_dir),
                str(infernal_bin_dir),
                str(minimap2_dir),
            ]
            + ([inherited_path] if inherited_path else [])
        ),
    }


def fetch_genbank(accession: str, out_path: Path, sleep_seconds: float, force: bool) -> None:
    if out_path.is_file() and out_path.stat().st_size > 0 and not force:
        return
    params = {
        "db": "nuccore",
        "id": accession,
        "rettype": "gbwithparts",
        "retmode": "text",
    }
    url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?" + urllib.parse.urlencode(params)
    request = urllib.request.Request(url, headers={"User-Agent": "metagenomic-astro-vadr-builder/1.0"})
    with urllib.request.urlopen(request, timeout=120) as response:
        text = response.read().decode("utf-8", errors="ignore")
    if not text.startswith("LOCUS"):
        raise RuntimeError(f"{accession}: invalid GenBank response")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(text, encoding="utf-8")
    time.sleep(sleep_seconds)


def run_command(cmd: list[str], env: dict[str, str]) -> None:
    subprocess.run(cmd, check=True, env=env)


def build_single_model(row: dict[str, str], genus_work_dir: Path, env: dict[str, str], force: bool) -> Path:
    accession = row["accession"]
    gb_path = genus_work_dir / "genbank" / f"{accession}.gb"
    model_dir = genus_work_dir / "models" / accession
    model_dir.parent.mkdir(parents=True, exist_ok=True)
    if model_dir.exists() and force:
        shutil.rmtree(model_dir)
    if model_dir.exists() and (model_dir / f"{accession}.vadr.minfo").is_file():
        return model_dir
    fasta_path = Path(row["fasta_path"])
    cmd = [
        str(NCOV_PERL),
        str((SOFT_DIR / "vadr" / "v-build.pl").resolve()),
        "-f",
        "--gb",
        "--ingb",
        str(gb_path),
        "--infa",
        str(fasta_path),
        "--group",
        GENUS_CONFIG[row["genus"]]["group"],
        "--subgroup",
        row["abbrev"] or accession,
        accession,
        str(model_dir),
    ]
    run_command(cmd, env)
    return model_dir


def concatenate_files(source_paths: list[Path], out_path: Path) -> None:
    with out_path.open("w", encoding="utf-8") as out_handle:
        for path in source_paths:
            out_handle.write(path.read_text(encoding="utf-8", errors="ignore"))


def build_genus_library(genus: str, rows: list[dict[str, str]], env: dict[str, str], sleep_seconds: float, force: bool) -> dict[str, str]:
    config = GENUS_CONFIG[genus]
    genus_key = config["model_key"]
    genus_work_dir = WORK_ROOT / genus_key
    library_dir = MODEL_ROOT / genus_key
    (genus_work_dir / "genbank").mkdir(parents=True, exist_ok=True)
    library_dir.mkdir(parents=True, exist_ok=True)

    success_rows: list[dict[str, str]] = []
    model_dirs: list[Path] = []
    failure_rows: list[dict[str, str]] = []
    for row in rows:
        try:
            fetch_genbank(row["accession"], genus_work_dir / "genbank" / f"{row['accession']}.gb", sleep_seconds, force)
            model_dir = build_single_model(row, genus_work_dir, env, force)
            success_rows.append(row)
            model_dirs.append(model_dir)
        except Exception as exc:  # noqa: BLE001
            failure_rows.append(
                {
                    "genus": row["genus"],
                    "species": row["species"],
                    "virus_name": row["virus_name"],
                    "isolate": row["isolate"],
                    "abbrev": row["abbrev"],
                    "accession": row["accession"],
                    "accession_full": row["accession_full"],
                    "reason": str(exc),
                }
            )

    if not success_rows:
        raise RuntimeError(f"{genus}: no models were built successfully")

    minfo_files = [model_dir / f"{row['accession']}.vadr.minfo" for row, model_dir in zip(success_rows, model_dirs)]
    cm_files = [model_dir / f"{row['accession']}.vadr.cm" for row, model_dir in zip(success_rows, model_dirs)]
    fa_files = [model_dir / f"{row['accession']}.vadr.fa" for row, model_dir in zip(success_rows, model_dirs)]
    hmm_files = [model_dir / f"{row['accession']}.vadr.protein.hmm" for row, model_dir in zip(success_rows, model_dirs)]

    out_minfo = library_dir / f"{genus_key}.minfo"
    out_cm = library_dir / f"{genus_key}.cm"
    out_fa = library_dir / f"{genus_key}.fa"
    out_hmm = library_dir / f"{genus_key}.hmm"

    concatenate_files(minfo_files, out_minfo)
    concatenate_files(cm_files, out_cm)
    concatenate_files(fa_files, out_fa)
    concatenate_files(hmm_files, out_hmm)

    # Copy per-model protein BLAST databases into the library root.
    for row, model_dir in zip(success_rows, model_dirs):
        accession = row["accession"]
        for path in model_dir.glob(f"{accession}.vadr.protein.fa*"):
            shutil.copy2(path, library_dir / path.name)

    run_command([str((SOFT_DIR / "infernal" / "binaries" / "cmpress").resolve()), str(out_cm)], env)
    run_command([str((SOFT_DIR / "infernal" / "binaries" / "hmmpress").resolve()), "-f", str(out_hmm)], env)
    run_command([str((SOFT_DIR / "infernal" / "binaries" / "esl-sfetch").resolve()), "--index", str(out_fa)], env)
    run_command(
        [
            str((SOFT_DIR / "ncbi-blast" / "bin" / "makeblastdb").resolve()),
            "-in",
            str(out_fa),
            "-dbtype",
            "nucl",
        ],
        env,
    )

    manifest_path = library_dir / f"{genus_key}.models.tsv"
    with manifest_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["genus", "species", "virus_name", "isolate", "abbrev", "accession", "accession_full", "fasta_path", "model_dir"],
            delimiter="\t",
        )
        writer.writeheader()
        for row, model_dir in zip(success_rows, model_dirs):
            writer.writerow(
                {
                    "genus": row["genus"],
                    "species": row["species"],
                    "virus_name": row["virus_name"],
                    "isolate": row["isolate"],
                    "abbrev": row["abbrev"],
                    "accession": row["accession"],
                    "accession_full": row["accession_full"],
                    "fasta_path": row["fasta_path"],
                    "model_dir": str(model_dir.resolve()),
                }
            )

    failure_path = library_dir / f"{genus_key}.failed.tsv"
    with failure_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["genus", "species", "virus_name", "isolate", "abbrev", "accession", "accession_full", "reason"],
            delimiter="\t",
        )
        writer.writeheader()
        for row in failure_rows:
            writer.writerow(row)

    return {
        "genus": genus,
        "model_key": genus_key,
        "library_dir": str(library_dir.resolve()),
        "models": str(len(success_rows)),
        "failed_models": str(len(failure_rows)),
        "manifest": str(manifest_path.resolve()),
        "failed_manifest": str(failure_path.resolve()),
    }


def main() -> int:
    args = parse_args()
    env = build_vadr_env()
    rows = load_complete_rows()
    selected_genera = ["Avastrovirus", "Mamastrovirus"] if args.genus == "all" else [args.genus]
    summary_rows: list[dict[str, str]] = []

    for genus in selected_genera:
        genus_rows = [row for row in rows if row["genus"] == genus]
        if not genus_rows:
            continue
        summary_rows.append(build_genus_library(genus, genus_rows, env, args.sleep_seconds, args.force))

    summary_path = WORK_ROOT / "build_summary.tsv"
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    with summary_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["genus", "model_key", "library_dir", "models", "failed_models", "manifest", "failed_manifest"],
            delimiter="\t",
        )
        writer.writeheader()
        for row in summary_rows:
            writer.writerow(row)

    print(f"Built {len(summary_rows)} astroviridae VADR model libraries")
    print(f"Summary: {summary_path.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
