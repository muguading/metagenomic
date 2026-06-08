#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import json
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/wuhhh/Desktop/徐老师/代码/metagenomic")
TYPING_XLSX = ROOT / "database" / "virus" / "enterovirus" / "typing.xlsx"
REF_DIR = ROOT / "database" / "virus" / "enterovirus" / "reference_genomes"
GFF_DIR = REF_DIR / "gff3"
OUT_DIR = REF_DIR / "abcd_vp1"
OUT_DIR.mkdir(parents=True, exist_ok=True)
EXPANDED_MANIFEST = OUT_DIR / "enterovirus_abcd_complete_genomes_expanded_manifest.tsv"
EXPANDED_FASTA = OUT_DIR / "enterovirus_abcd_complete_genomes_expanded.fasta"
REPORT_TSV = OUT_DIR / "enterovirus_abcd_complete_genomes_expanded_download_report.tsv"

ESEARCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
ESUMMARY_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi"
EFETCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
GFF_URL = "https://www.ncbi.nlm.nih.gov/sviewer/viewer.fcgi"
MAX_WORKERS = 6
REQUEST_SLEEP = 0.34

SPECIES_TO_GROUP = {
    "Enterovirus alphacoxsackie": "A",
    "Enterovirus betacoxsackie": "B",
    "Enterovirus coxsackiepol": "C",
    "Enterovirus d": "D",
    "Enterovirus alpharhino": "A",
    "Enterovirus betarhino": "B",
    "Enterovirus cerhino": "C",
}


def fetch_text(url: str, params: dict[str, str], retries: int = 6, timeout: int = 60) -> str:
    query = urllib.parse.urlencode(params)
    full_url = f"{url}?{query}"
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            request = urllib.request.Request(
                full_url,
                headers={"User-Agent": "metagenomic-enterovirus-abcd-expanded/1.0"},
            )
            with urllib.request.urlopen(request, timeout=timeout) as response:
                return response.read().decode("utf-8", errors="replace")
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            if attempt == retries:
                break
            delay = max(1.5, attempt * 2.0)
            if isinstance(exc, urllib.error.HTTPError) and exc.code == 429:
                delay = max(delay, 6.0 * attempt)
            time.sleep(delay)
    raise RuntimeError(f"failed to fetch {full_url}: {last_error}") from last_error


def fetch_json(url: str, params: dict[str, str], timeout: int = 60) -> dict:
    return json.loads(fetch_text(url, params, timeout=timeout))


def load_typing_rows() -> list[dict[str, str]]:
    wb = load_workbook(TYPING_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(name or "").strip() for name in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    result: list[dict[str, str]] = []
    for values in rows[1:]:
        accession = str(values[idx["Accession"]] or "").strip()
        if not accession or accession.startswith("partial:"):
            continue
        species = str(values[idx["Species"]] or "").strip()
        big_group = SPECIES_TO_GROUP.get(species, "")
        if big_group not in {"A", "B", "C", "D"}:
            continue
        result.append(
            {
                "genus": str(values[idx["Genus"]] or "").strip(),
                "species": species,
                "virus_name": str(values[idx["Virus Name"]] or "").strip(),
                "isolate": str(values[idx["Isolate"]] or "").strip(),
                "accession": accession.split(".", 1)[0],
                "accession_full": accession,
                "available": str(values[idx["Available"]] or "").strip(),
                "abbrev": str(values[idx["Abbrev."]] or "").strip(),
                "big_group": big_group,
            }
        )
    return result


def load_existing_manifest_rows() -> list[dict[str, str]]:
    if not EXPANDED_MANIFEST.is_file():
        return []
    with EXPANDED_MANIFEST.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def build_queries(row: dict[str, str]) -> list[str]:
    virus_name = row["virus_name"].strip()
    abbrev = row["abbrev"].strip()
    queries: list[str] = []
    if virus_name:
        queries.append(f'"{virus_name}"[Title] AND ("complete genome"[Title] OR "complete sequence"[Title])')
        queries.append(f'"{virus_name}"[All Fields] AND ("complete genome"[Title] OR "complete sequence"[Title])')
    if abbrev.startswith("RV-"):
        suffix = abbrev.replace("RV-", "")
        queries.append(f'"rhinovirus {suffix}"[Title] AND ("complete genome"[Title] OR "complete sequence"[Title])')
        queries.append(f'"human rhinovirus {suffix}"[Title] AND ("complete genome"[Title] OR "complete sequence"[Title])')
    elif abbrev.startswith("EV-"):
        suffix = abbrev.replace("EV-", "")
        queries.append(f'"enterovirus {suffix}"[Title] AND ("complete genome"[Title] OR "complete sequence"[Title])')
    elif abbrev.startswith("CVA"):
        suffix = abbrev.replace("CVA", "")
        queries.append(f'"coxsackievirus A{suffix}"[Title] AND ("complete genome"[Title] OR "complete sequence"[Title])')
    elif abbrev.startswith("CVB"):
        suffix = abbrev.replace("CVB", "")
        queries.append(f'"coxsackievirus B{suffix}"[Title] AND ("complete genome"[Title] OR "complete sequence"[Title])')
    elif abbrev.startswith("E") and abbrev[1:].isdigit():
        queries.append(f'"echovirus {abbrev[1:]}"[Title] AND ("complete genome"[Title] OR "complete sequence"[Title])')
    return list(dict.fromkeys(query for query in queries if query))


def esearch_ids(term: str, retmax: int = 500) -> list[str]:
    payload = fetch_json(
        ESEARCH_URL,
        {
            "db": "nuccore",
            "retmode": "json",
            "term": term,
            "retmax": str(retmax),
        },
    )
    time.sleep(REQUEST_SLEEP)
    return payload.get("esearchresult", {}).get("idlist", [])


def esummary_many(uids: list[str]) -> list[dict[str, str]]:
    if not uids:
        return []
    payload = fetch_json(
        ESUMMARY_URL,
        {"db": "nuccore", "id": ",".join(uids), "retmode": "json"},
    )
    time.sleep(REQUEST_SLEEP)
    result = payload.get("result", {})
    ordered = []
    for uid in uids:
        row = result.get(uid)
        if isinstance(row, dict):
            ordered.append(row)
    return ordered


def looks_complete(title: str, length: int) -> bool:
    text = (title or "").lower()
    if "nearly complete" in text or "near complete" in text or "partial" in text:
        return False
    if "complete genome" in text:
        return True
    if "complete sequence" in text and length >= 6000:
        return True
    if "complete cds" in text and length >= 6000:
        return True
    return False


def subtype_hits_title(row: dict[str, str], title: str) -> bool:
    text = (title or "").lower()
    abbrev = row["abbrev"].strip().lower()
    virus_name = row["virus_name"].strip().lower()
    if virus_name and virus_name in text:
        return True
    if abbrev.startswith("RV-"):
        suffix = abbrev.replace("rv-", "")
        return f"rhinovirus {suffix}" in text
    if abbrev.startswith("EV-"):
        suffix = abbrev.replace("ev-", "")
        return f"enterovirus {suffix}" in text
    if abbrev.startswith("CVA"):
        return f"coxsackievirus a{abbrev[3:].lower()}" in text
    if abbrev.startswith("CVB"):
        return f"coxsackievirus b{abbrev[3:].lower()}" in text
    if abbrev.startswith("E") and abbrev[1:].isdigit():
        return f"echovirus {abbrev[1:]}" in text
    return False


def discover_complete_genomes_for_subtype(row: dict[str, str]) -> list[dict[str, str]]:
    found: dict[str, dict[str, str]] = {}
    for term in build_queries(row):
        uids = esearch_ids(term)
        if not uids:
            continue
        for index in range(0, len(uids), 50):
            chunk = uids[index:index + 50]
            for summary in esummary_many(chunk):
                accession_full = str(summary.get("caption") or "").strip()
                accession = accession_full.split(".", 1)[0] if accession_full else ""
                title = str(summary.get("title") or "").strip()
                try:
                    length = int(summary.get("slen") or summary.get("length") or 0)
                except (TypeError, ValueError):
                    length = 0
                if not accession:
                    continue
                if not looks_complete(title, length):
                    continue
                if not subtype_hits_title(row, title):
                    continue
                found.setdefault(
                    accession,
                    {
                        "genus": row["genus"],
                        "species": row["species"],
                        "virus_name": row["virus_name"],
                        "isolate": title,
                        "accession": accession,
                        "accession_full": accession_full or accession,
                        "available": "Complete genome",
                        "abbrev": row["abbrev"],
                        "big_group": row["big_group"],
                        "title": title,
                        "sequence_length": str(length),
                        "source_query": term,
                    },
                )
    return sorted(found.values(), key=lambda item: item["accession"])


def download_fasta(accession: str, out_path: Path) -> str:
    if out_path.is_file() and out_path.stat().st_size > 0:
        return "cached"
    fasta_text = fetch_text(
        EFETCH_URL,
        {"db": "nuccore", "id": accession, "rettype": "fasta", "retmode": "text"},
    )
    if not fasta_text.lstrip().startswith(">"):
        raise RuntimeError(f"{accession}: invalid FASTA response")
    out_path.write_text(fasta_text, encoding="utf-8")
    time.sleep(REQUEST_SLEEP)
    return "downloaded"


def download_gff(accession: str, out_path: Path) -> str:
    if out_path.is_file() and out_path.stat().st_size > 0:
        return "cached"
    gff_text = fetch_text(
        GFF_URL,
        {"id": accession, "db": "nuccore", "report": "gff3", "retmode": "text"},
    )
    if "##gff-version" not in gff_text:
        raise RuntimeError(f"{accession}: invalid GFF3 response")
    out_path.write_text(gff_text, encoding="utf-8")
    time.sleep(REQUEST_SLEEP)
    return "downloaded"


def fetch_one_reference(row: dict[str, str]) -> dict[str, str]:
    accession = row["accession"]
    fasta_path = REF_DIR / f"{accession}.fasta"
    gff_path = GFF_DIR / f"{accession}.gff3"
    fasta_path.parent.mkdir(parents=True, exist_ok=True)
    gff_path.parent.mkdir(parents=True, exist_ok=True)
    fasta_status = download_fasta(accession, fasta_path)
    gff_status = download_gff(accession, gff_path)
    header = ""
    sequence_length = row.get("sequence_length", "")
    with fasta_path.open("r", encoding="utf-8", errors="ignore") as handle:
        first = handle.readline().strip()
        header = first[1:] if first.startswith(">") else first
        if not sequence_length:
            seq = "".join(line.strip() for line in handle if line and not line.startswith(">"))
            sequence_length = str(len(seq))
    result = dict(row)
    result.update(
        {
            "header": header,
            "sequence_length": sequence_length,
            "fasta_path": str(fasta_path),
            "gff_path": str(gff_path),
            "fasta_status": fasta_status,
            "gff_status": gff_status,
            "status": "ok",
        }
    )
    return result


def write_outputs(results: list[dict[str, str]], report_rows: list[dict[str, str]]) -> None:
    fieldnames = [
        "genus",
        "species",
        "virus_name",
        "isolate",
        "accession",
        "accession_full",
        "available",
        "abbrev",
        "big_group",
        "title",
        "header",
        "sequence_length",
        "fasta_path",
        "gff_path",
        "fasta_status",
        "gff_status",
        "source_query",
        "status",
    ]
    with EXPANDED_MANIFEST.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in results:
            writer.writerow({key: row.get(key, "") for key in fieldnames})

    with EXPANDED_FASTA.open("w", encoding="utf-8") as out_handle:
        for row in results:
            fasta_path = Path(row["fasta_path"])
            if fasta_path.is_file():
                out_handle.write(fasta_path.read_text(encoding="utf-8", errors="ignore").rstrip() + "\n")

    report_fields = [
        "abbrev",
        "big_group",
        "virus_name",
        "status",
        "candidate_count",
        "downloaded_count",
        "note",
    ]
    with REPORT_TSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=report_fields, delimiter="\t")
        writer.writeheader()
        for row in report_rows:
            writer.writerow({key: row.get(key, "") for key in report_fields})


def main() -> int:
    typing_rows = load_typing_rows()
    subtype_meta: dict[str, dict[str, str]] = {}
    for row in typing_rows:
        subtype_meta.setdefault(row["abbrev"], row)

    existing_rows = load_existing_manifest_rows()
    combined_by_accession = {str(row.get("accession") or "").strip(): row for row in existing_rows if str(row.get("accession") or "").strip()}
    report_rows: list[dict[str, str]] = []

    for abbrev, row in sorted(subtype_meta.items(), key=lambda item: (item[1]["big_group"], item[0])):
        try:
            candidates = discover_complete_genomes_for_subtype(row)
            downloaded_count = 0
            for candidate in candidates:
                accession = candidate["accession"]
                previous = combined_by_accession.get(accession, {})
                merged = dict(previous)
                merged.update(candidate)
                combined_by_accession[accession] = merged
                downloaded_count += 1
            report_rows.append(
                {
                    "abbrev": abbrev,
                    "big_group": row["big_group"],
                    "virus_name": row["virus_name"],
                    "status": "ok",
                    "candidate_count": str(len(candidates)),
                    "downloaded_count": str(downloaded_count),
                    "note": "",
                }
            )
            print(f"[discover] {abbrev}\t{len(candidates)}")
        except Exception as exc:  # noqa: BLE001
            report_rows.append(
                {
                    "abbrev": abbrev,
                    "big_group": row["big_group"],
                    "virus_name": row["virus_name"],
                    "status": "error",
                    "candidate_count": "0",
                    "downloaded_count": "0",
                    "note": str(exc),
                }
            )
            print(f"[discover-fail] {abbrev}\t{exc}")

    pending_rows = sorted(
        [row for row in combined_by_accession.values() if str(row.get("accession") or "").strip()],
        key=lambda item: (item.get("big_group", ""), item.get("abbrev", ""), item.get("accession", "")),
    )

    downloaded_results: list[dict[str, str]] = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map = {executor.submit(fetch_one_reference, row): row for row in pending_rows}
        for future in as_completed(future_map):
            row = future_map[future]
            accession = row["accession"]
            try:
                result = future.result()
                downloaded_results.append(result)
                print(f"[ok] {row['abbrev']}\t{accession}")
            except Exception as exc:  # noqa: BLE001
                failed = dict(row)
                failed["status"] = "error"
                failed["note"] = str(exc)
                downloaded_results.append(failed)
                print(f"[fail] {row['abbrev']}\t{accession}\t{exc}")

    downloaded_results.sort(key=lambda item: (item.get("big_group", ""), item.get("abbrev", ""), item.get("accession", "")))
    write_outputs(downloaded_results, report_rows)
    ok_count = sum(1 for row in downloaded_results if row.get("status") == "ok")
    print(f"Expanded complete genomes written: {ok_count}")
    print(f"Manifest: {EXPANDED_MANIFEST}")
    print(f"Report: {REPORT_TSV}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
