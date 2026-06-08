#!/usr/bin/env python3
from __future__ import annotations

import csv
import os
import shlex
import subprocess
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/wuhhh/Desktop/徐老师/代码/metagenomic")
REFERENCE_DIR = ROOT / "database" / "virus" / "enterovirus" / "reference_genomes"
ORIGINAL_GFF_DIR = REFERENCE_DIR / "gff3"
VADR_GFF_DIR = REFERENCE_DIR / "gff3_vadr"
VADR_WORK_DIR = REFERENCE_DIR / "vadr_annotation"
MISSING_TSV = REFERENCE_DIR / "enterovirus_vp1_missing.tsv"
STATUS_TSV = VADR_GFF_DIR / "annotation_status.tsv"
TYPING_XLSX = ROOT / "database" / "virus" / "enterovirus" / "typing.xlsx"

SPECIES_TO_MODEL = {
    "Enterovirus alphacoxsackie": "evA",
    "Enterovirus alpharhino": "evA",
    "Enterovirus betacoxsackie": "evB",
    "Enterovirus betarhino": "evB",
    "Enterovirus cerhino": "evC",
    "Enterovirus coxsackiepol": "evC",
    "Enterovirus deconjuncti": "evD",
}


def parse_fasta(path: Path) -> dict[str, str]:
    records: dict[str, str] = {}
    header: str | None = None
    chunks: list[str] = []
    with path.open() as handle:
        for raw_line in handle:
            line = raw_line.rstrip("\n")
            if line.startswith(">"):
                if header is not None:
                    records[header] = "".join(chunks)
                header = line[1:]
                chunks = []
            else:
                chunks.append(line)
    if header is not None:
        records[header] = "".join(chunks)
    return records


def load_typing_meta() -> dict[str, dict[str, str]]:
    wb = load_workbook(TYPING_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {str(name): i for i, name in enumerate(header)}
    meta: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        accession = str(row[idx["Accession"]] or "").strip()
        if not accession or accession.startswith("partial:"):
            continue
        accession_root = accession.split(".", 1)[0]
        meta[accession_root] = {
            "accession": accession,
            "genus": str(row[idx["Genus"]] or "").strip(),
            "species": str(row[idx["Species"]] or "").strip(),
            "virus_name": str(row[idx["Virus Name"]] or "").strip(),
            "isolate": str(row[idx["Isolate"]] or "").strip(),
            "abbrev": str(row[idx["Abbrev."]] or "").strip(),
            "available": str(row[idx["Available"]] or "").strip(),
        }
    return meta


def build_vadr_env(model_dir: Path) -> dict[str, str]:
    vadr_root = (ROOT / "soft").resolve()
    vadr_scripts_dir = (vadr_root / "vadr").resolve()
    infernal_bin_dir = (vadr_root / "infernal" / "binaries").resolve()
    bio_easel_dir = (vadr_root / "Bio-Easel").resolve()
    alt_bio_easel_dir = (vadr_root / "Bio-Easel-ncov").resolve()
    if (alt_bio_easel_dir / "blib" / "lib").exists() and (alt_bio_easel_dir / "blib" / "arch").exists():
        bio_easel_dir = alt_bio_easel_dir
    sequip_dir = (vadr_root / "sequip").resolve()
    blast_bin_dir = (vadr_root / "ncbi-blast" / "bin").resolve()
    fasta_bin_dir = (vadr_root / "fasta" / "bin").resolve()
    minimap2_dir = (vadr_root / "minimap2").resolve()
    required_paths = [
        vadr_scripts_dir / "v-annotate.pl",
        vadr_scripts_dir / "miniscripts" / "annotate-tbl2gff.pl",
        vadr_scripts_dir / "miniscripts" / "fasta-trim-terminal-ambigs.pl",
        infernal_bin_dir / "cmalign",
        bio_easel_dir / "blib" / "lib",
        bio_easel_dir / "blib" / "arch",
        sequip_dir,
        blast_bin_dir / "blastn",
        fasta_bin_dir / "fasta36",
        minimap2_dir / "minimap2",
        model_dir,
    ]
    missing = [str(path) for path in required_paths if not path.exists()]
    if missing:
        raise FileNotFoundError("VADR dependencies missing: " + ", ".join(missing[:10]))
    inherited_path = str(os.environ.get("PATH") or "").strip()
    inherited_perl5lib = str(os.environ.get("PERL5LIB") or "").strip()
    return {
        **os.environ,
        "VADRINSTALLDIR": str(vadr_root),
        "VADRSCRIPTSDIR": str(vadr_scripts_dir),
        "VADRCONFIGFILE": str((vadr_scripts_dir / "vadr.config").resolve()),
        "VADRMODELDIR": str(model_dir.resolve()),
        "VADRINFERNALDIR": str(infernal_bin_dir),
        "VADREASELDIR": str(infernal_bin_dir),
        "VADRHMMERDIR": str(infernal_bin_dir),
        "VADRBIOEASELDIR": str(bio_easel_dir),
        "VADRSEQUIPDIR": str(sequip_dir),
        "VADRBLASTDIR": str(blast_bin_dir),
        "VADRFASTADIR": str(fasta_bin_dir),
        "VADRMINIMAP2DIR": str(minimap2_dir),
        "PERL5LIB": os.pathsep.join(
            [
                str(vadr_scripts_dir),
                str(sequip_dir),
                str((bio_easel_dir / "blib" / "lib").resolve()),
                str((bio_easel_dir / "blib" / "arch").resolve()),
            ]
            + ([inherited_perl5lib] if inherited_perl5lib else [])
        ),
        "PATH": os.pathsep.join(
            [
                str(vadr_scripts_dir),
                str(blast_bin_dir),
                str(fasta_bin_dir),
                str(infernal_bin_dir),
                str(minimap2_dir),
            ]
            + ([inherited_path] if inherited_path else [])
        ),
    }


def resolve_vadr_perl_bin() -> str:
    for candidate in ["/usr/bin/perl", "perl"]:
        try:
            completed = subprocess.run(
                [candidate, "-MInline", "-e", "print qq(ok\\n)"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                check=False,
            )
            if completed.returncode == 0:
                return candidate
        except Exception:
            continue
    return "perl"


def resolve_working_perl_with_module(env: dict[str, str] | None, module_name: str, candidates: list[str]) -> str:
    for candidate in candidates:
        if not candidate:
            continue
        try:
            completed = subprocess.run(
                [candidate, f"-M{module_name}", "-e", "print qq(ok\\n)"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                check=False,
                env=env,
            )
            if completed.returncode == 0:
                return candidate
        except Exception:
            continue
    return candidates[0] if candidates else "perl"


def run_command(cmd: str, env: dict[str, str], cwd: Path | None = None) -> None:
    subprocess.run(cmd, shell=True, check=True, env=env, cwd=str(cwd) if cwd else None)


def split_group_gff(group_gff: Path, output_map: dict[str, Path]) -> None:
    buckets: dict[str, list[str]] = {seqid: [] for seqid in output_map}
    current_headers: list[str] = []
    with group_gff.open() as handle:
        for raw_line in handle:
            if raw_line.startswith("#"):
                current_headers.append(raw_line)
                continue
            fields = raw_line.rstrip("\n").split("\t")
            if len(fields) != 9:
                continue
            seqid = fields[0]
            if seqid not in buckets:
                continue
            if not buckets[seqid]:
                buckets[seqid].extend(current_headers)
            buckets[seqid].append(raw_line)
    for seqid, out_path in output_map.items():
        content = buckets.get(seqid, [])
        if not content:
            continue
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with out_path.open("w") as handle:
            handle.writelines(content)


def tbl_has_feature_rows(tbl_path: Path) -> bool:
    if not tbl_path.exists():
        return False
    with tbl_path.open() as handle:
        for raw_line in handle:
            if raw_line.startswith(">Feature "):
                return True
    return False


def main() -> None:
    VADR_GFF_DIR.mkdir(parents=True, exist_ok=True)
    VADR_WORK_DIR.mkdir(parents=True, exist_ok=True)

    typing_meta = load_typing_meta()
    genomes = parse_fasta(REFERENCE_DIR / "enterovirus_typing_complete_genomes.fasta")

    with MISSING_TSV.open() as handle:
        missing_rows = list(csv.DictReader(handle, delimiter="\t"))

    grouped_rows: dict[str, list[dict[str, str]]] = defaultdict(list)
    status_rows: list[dict[str, str]] = []

    for row in missing_rows:
        accession_root = row["accession"]
        meta = typing_meta.get(accession_root)
        if not meta:
            status_rows.append(
                {
                    "accession_root": accession_root,
                    "accession": accession_root,
                    "species": "",
                    "virus_name": "",
                    "abbrev": "",
                    "model_key": "",
                    "status": "missing_typing_metadata",
                    "note": "typing.xlsx 中未找到对应 accession",
                    "vadr_gff_path": "",
                }
            )
            continue
        model_key = SPECIES_TO_MODEL.get(meta["species"], "")
        if not model_key:
            status_rows.append(
                {
                    "accession_root": accession_root,
                    "accession": meta["accession"],
                    "species": meta["species"],
                    "virus_name": meta["virus_name"],
                    "abbrev": meta["abbrev"],
                    "model_key": "",
                    "status": "unsupported_species",
                    "note": "当前 vadr-models-ev 仅覆盖 evA/evB/evC/evD",
                    "vadr_gff_path": "",
                }
            )
            continue
        grouped_rows[model_key].append({**row, **meta, "accession_root": accession_root, "model_key": model_key})

    for model_key, group_rows in grouped_rows.items():
        model_dir = ROOT / "soft" / "vadr-models-ev" / model_key
        env = build_vadr_env(model_dir)
        ncov_bin_dir = "/opt/homebrew/Caskroom/mambaforge/base/envs/ncov/bin"
        env["PATH"] = os.pathsep.join([ncov_bin_dir, "/usr/bin", "/bin", str(env.get("PATH") or "")])
        perl_candidates = [f"{ncov_bin_dir}/perl", "/usr/bin/perl", resolve_vadr_perl_bin(), "perl"]
        perl_bin = resolve_working_perl_with_module(env, "Bio::Easel::MSA", perl_candidates)
        group_dir = VADR_WORK_DIR / model_key
        group_dir.mkdir(parents=True, exist_ok=True)
        input_fasta = group_dir / f"{model_key}.input.fasta"
        trimmed_fasta = group_dir / f"{model_key}.trimmed.fasta"
        output_dir = group_dir / f"{model_key}_vadr"
        prefix = f"{output_dir.name}.vadr"
        pass_tbl = output_dir / f"{prefix}.pass.tbl"
        fail_tbl = output_dir / f"{prefix}.fail.tbl"
        merged_gff = group_dir / f"{model_key}.vadr.gff3"

        with input_fasta.open("w") as handle:
            for row in group_rows:
                fasta_path = REFERENCE_DIR / f"{row['accession_root']}.fasta"
                if not fasta_path.exists():
                    continue
                records = parse_fasta(fasta_path)
                if not records:
                    continue
                header, seq = next(iter(records.items()))
                handle.write(f">{row['accession_root']}\n")
                for i in range(0, len(seq), 80):
                    handle.write(seq[i : i + 80] + "\n")

        trim_script = ROOT / "soft" / "vadr" / "miniscripts" / "fasta-trim-terminal-ambigs.pl"
        run_command(
            f"{shlex.quote(str(perl_bin))} {shlex.quote(str(trim_script))} --minlen 50 --maxlen 8000 {shlex.quote(str(input_fasta))} > {shlex.quote(str(trimmed_fasta))}",
            env=env,
        )

        vadr_script = ROOT / "soft" / "vadr" / "v-annotate.pl"
        cmd_parts = [
            shlex.quote(str(perl_bin)),
            shlex.quote(str(vadr_script)),
            "-f",
            "-r",
            "--ignore_exc",
            "--mkey",
            model_key,
            "--mdir",
            shlex.quote(str(model_dir)),
            shlex.quote(str(trimmed_fasta)),
            shlex.quote(str(output_dir)),
        ]
        run_command(" ".join(cmd_parts), env=env)

        annotate_tbl2gff = ROOT / "soft" / "vadr" / "miniscripts" / "annotate-tbl2gff.pl"
        source_tbl = pass_tbl if tbl_has_feature_rows(pass_tbl) else fail_tbl
        if not tbl_has_feature_rows(source_tbl):
            raise FileNotFoundError(f"No usable VADR feature table found for {model_key}: {source_tbl}")
        run_command(
            f"{shlex.quote(str(perl_bin))} {shlex.quote(str(annotate_tbl2gff))} {shlex.quote(str(source_tbl))} > {shlex.quote(str(merged_gff))}",
            env=env,
        )

        output_map = {row["accession_root"]: VADR_GFF_DIR / f"{row['accession_root']}.gff3" for row in group_rows}
        split_group_gff(merged_gff, output_map)

        failed_accessions: set[str] = set()
        if fail_tbl.exists():
            with fail_tbl.open() as handle:
                for raw_line in handle:
                    line = raw_line.strip()
                    if not line or line.startswith(">"):
                        continue
                    failed_accessions.add(line.split("\t", 1)[0].split()[0])

        for row in group_rows:
            out_gff = output_map[row["accession_root"]]
            status_rows.append(
                {
                    "accession_root": row["accession_root"],
                    "accession": row["accession"],
                    "species": row["species"],
                    "virus_name": row["virus_name"],
                    "abbrev": row["abbrev"],
                    "model_key": model_key,
                    "status": "annotated" if out_gff.exists() else ("failed" if row["accession_root"] in failed_accessions else "missing"),
                    "note": "",
                    "vadr_gff_path": str(out_gff),
                }
            )

    with STATUS_TSV.open("w", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["accession_root", "accession", "species", "virus_name", "abbrev", "model_key", "status", "note", "vadr_gff_path"],
            delimiter="\t",
        )
        writer.writeheader()
        writer.writerows(status_rows)


if __name__ == "__main__":
    main()
