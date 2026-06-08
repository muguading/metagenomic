#!/usr/bin/env python3
from __future__ import annotations

import csv
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/wuhhh/Desktop/徐老师/代码/metagenomic")
XLSX_PATH = ROOT / "database" / "virus" / "astroviridae" / "typing.xlsx"
OUT_DIR = ROOT / "database" / "virus" / "astroviridae" / "reference_genomes"
GFF_DIR = OUT_DIR / "gff3"
FASTA_MERGED = OUT_DIR / "astroviridae_typing_reference_genomes.fasta"
MANIFEST_PATH = OUT_DIR / "astroviridae_typing_reference_genomes_manifest.tsv"

FASTA_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
GFF_URL = "https://www.ncbi.nlm.nih.gov/sviewer/viewer.fcgi"
MAX_WORKERS = 3
REQUEST_SLEEP = 0.4


def fetch_text(url: str, params: dict[str, str], retries: int = 6, timeout: int = 60) -> str:
    query = urllib.parse.urlencode(params)
    full_url = f"{url}?{query}"
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            request = urllib.request.Request(
                full_url,
                headers={"User-Agent": "metagenomic-astroviridae-downloader/1.0"},
            )
            with urllib.request.urlopen(request, timeout=timeout) as response:
                data = response.read()
            return data.decode("utf-8", errors="replace")
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            if attempt == retries:
                break
            delay = max(1.5, attempt * 2.0)
            if isinstance(exc, urllib.error.HTTPError) and exc.code == 429:
                delay = max(delay, 6.0 * attempt)
            time.sleep(delay)
    raise RuntimeError(f"failed to fetch {full_url}: {last_error}") from last_error


def load_rows() -> list[dict[str, str]]:
    workbook = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        if not values or not values[4]:
            continue
        rows.append(
            {
                "genus": str(values[0] or "").strip(),
                "species": str(values[1] or "").strip(),
                "virus_name": str(values[2] or "").strip(),
                "isolate": str(values[3] or "").strip(),
                "accession": str(values[4] or "").strip(),
                "available_sequence": str(values[5] or "").strip(),
                "abbrev": str(values[6] or "").strip(),
            }
        )
    return rows


def fasta_target(accession: str) -> Path:
    return OUT_DIR / f"{accession}.fasta"


def gff_target(accession: str) -> Path:
    return GFF_DIR / f"{accession}.gff3"


def download_one(row: dict[str, str]) -> dict[str, str]:
    accession = row["accession"]
    fasta_path = fasta_target(accession)
    gff_path = gff_target(accession)
    fasta_path.parent.mkdir(parents=True, exist_ok=True)
    gff_path.parent.mkdir(parents=True, exist_ok=True)

    fasta_status = "cached"
    gff_status = "cached"

    if not fasta_path.is_file() or fasta_path.stat().st_size == 0:
        fasta_text = fetch_text(
            FASTA_URL,
            {
                "db": "nuccore",
                "id": accession,
                "rettype": "fasta",
                "retmode": "text",
            },
        )
        if not fasta_text.lstrip().startswith(">"):
            raise RuntimeError(f"{accession}: invalid FASTA response")
        fasta_path.write_text(fasta_text, encoding="utf-8")
        fasta_status = "downloaded"
        time.sleep(REQUEST_SLEEP)

    if not gff_path.is_file() or gff_path.stat().st_size == 0:
        gff_text = fetch_text(
            GFF_URL,
            {
                "id": accession,
                "db": "nuccore",
                "report": "gff3",
                "retmode": "text",
            },
        )
        if "##gff-version" not in gff_text:
            raise RuntimeError(f"{accession}: invalid GFF response")
        gff_path.write_text(gff_text, encoding="utf-8")
        gff_status = "downloaded"
        time.sleep(REQUEST_SLEEP)

    header = ""
    sequence_length = ""
    with fasta_path.open("r", encoding="utf-8", errors="ignore") as handle:
        first_line = handle.readline().strip()
        header = first_line[1:] if first_line.startswith(">") else first_line
        sequence = "".join(line.strip() for line in handle if line and not line.startswith(">"))
        sequence_length = str(len(sequence))

    result = dict(row)
    result.update(
        {
            "header": header,
            "sequence_length": sequence_length,
            "fasta_path": str(fasta_path),
            "gff_path": str(gff_path),
            "fasta_status": fasta_status,
            "gff_status": gff_status,
            "status": "ok",
        }
    )
    return result


def write_outputs(results: list[dict[str, str]]) -> None:
    with FASTA_MERGED.open("w", encoding="utf-8") as out_handle:
        for row in results:
            fasta_path = Path(row["fasta_path"])
            out_handle.write(fasta_path.read_text(encoding="utf-8", errors="ignore").rstrip() + "\n")

    fieldnames = [
        "genus",
        "species",
        "virus_name",
        "isolate",
        "accession",
        "available_sequence",
        "abbrev",
        "header",
        "sequence_length",
        "fasta_path",
        "gff_path",
        "fasta_status",
        "gff_status",
        "status",
    ]
    with MANIFEST_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in results:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def main() -> int:
    rows = load_rows()
    if not rows:
        print("No accession rows found in typing.xlsx", file=sys.stderr)
        return 1

    results: list[dict[str, str]] = []
    errors: list[tuple[str, str]] = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map = {executor.submit(download_one, row): row["accession"] for row in rows}
        for future in as_completed(future_map):
            accession = future_map[future]
            try:
                result = future.result()
                results.append(result)
                print(f"[ok] {accession}")
            except Exception as exc:  # noqa: BLE001
                errors.append((accession, str(exc)))
                print(f"[fail] {accession}: {exc}", file=sys.stderr)

    results.sort(key=lambda item: item["accession"])
    write_outputs(results)

    if errors:
        print(f"{len(errors)} accessions failed", file=sys.stderr)
        for accession, message in errors[:20]:
            print(f"  - {accession}: {message}", file=sys.stderr)
        return 2

    print(f"Downloaded {len(results)} astroviridae references to {OUT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
