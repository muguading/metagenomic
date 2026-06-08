#!/usr/bin/env python3
from __future__ import annotations

import csv
import re
from pathlib import Path
from urllib.parse import unquote

from openpyxl import load_workbook


ROOT = Path("/Users/wuhhh/Desktop/徐老师/代码/metagenomic")
REF_DIR = ROOT / "database" / "virus" / "astroviridae" / "reference_genomes"
TYPING_XLSX = ROOT / "database" / "virus" / "astroviridae" / "typing.xlsx"
GFF_DIR = REF_DIR / "gff3"
OUT_DIR = REF_DIR / "orf2_capsid"
MERGED_FASTA = REF_DIR / "astroviridae_orf2_capsid_references.fasta"
MANIFEST_OUT = REF_DIR / "astroviridae_orf2_capsid_references.tsv"
MISSING_MANIFEST = REF_DIR / "astroviridae_orf2_capsid_missing.tsv"


def parse_attributes(text: str) -> dict[str, str]:
    attrs: dict[str, str] = {}
    for item in str(text or "").strip().split(";"):
        if not item:
            continue
        if "=" not in item:
            continue
        key, value = item.split("=", 1)
        attrs[key] = value
    return attrs


def normalize_label(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", unquote(str(text or "")).strip().lower())


def is_orf2_capsid_annotation(product: str, gene: str, note: str) -> bool:
    raw_values = [str(product or "").strip().lower(), str(gene or "").strip().lower(), str(note or "").strip().lower()]
    normalized = [normalize_label(value) for value in raw_values]
    product_l = raw_values[0]
    product_n = normalized[0]

    if any("orf2" in value or "orf 2" in value for value in raw_values):
        return True
    if any(value in {"orf2", "orftwo", "ofr2"} for value in normalized):
        return True
    if any("capsid" in value for value in raw_values):
        return True
    if ("structural polyprotein" in product_l or product_n == "structuralpolyprotein") and "non-structural" not in product_l:
        return True
    return False


def score_orf2_capsid_annotation(product: str, gene: str, note: str) -> int:
    product_l = str(product or "").strip().lower()
    gene_l = str(gene or "").strip().lower()
    note_l = str(note or "").strip().lower()
    product_n = normalize_label(product_l)
    gene_n = normalize_label(gene_l)
    note_n = normalize_label(note_l)

    score = 0
    if gene_n in {"orf2", "ofr2"}:
        score += 10
    if "orf2" in product_l or "orf 2" in product_l or product_n == "ofr2":
        score += 8
    if "orf2" in note_l or "orf 2" in note_l:
        score += 4
    if "capsid" in product_l:
        score += 8
    elif "capsid" in note_l:
        score += 4
    if "structural polyprotein" in product_l or "structuralpolyprotein" in product_n:
        score += 7

    if "non-structural" in product_l or "nonstructural" in product_n:
        score -= 8
    if gene_n.startswith("orf1") or "orf1" in product_n or "orf1" in note_n:
        score -= 6
    if "polyprotein1ab" in product_n or "rna-dependentrnapolymerase" in note_n:
        score -= 4
    return score


def find_orf2_capsid_feature(gff_path: Path) -> dict[str, str] | None:
    best: dict[str, str] | None = None
    with gff_path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line in handle:
            if not line or line.startswith("#"):
                continue
            parts = line.rstrip("\n").split("\t")
            if len(parts) != 9:
                continue
            seqid, source, feature_type, start, end, score, strand, phase, attributes = parts
            attrs = parse_attributes(attributes)
            product = str(attrs.get("product") or "").strip()
            gene = str(attrs.get("gene") or "").strip()
            note = str(attrs.get("Note") or attrs.get("note") or "").strip()
            if not is_orf2_capsid_annotation(product, gene, note):
                continue
            score = score_orf2_capsid_annotation(product, gene, note)
            if score <= 0:
                continue
            candidate = {
                "seqid": seqid,
                "feature_type": feature_type,
                "start": start,
                "end": end,
                "strand": strand,
                "product": product,
                "gene": gene,
                "note": note,
                "partial": str(attrs.get("partial") or ""),
                "score": str(score),
            }
            if best is None:
                best = candidate
                continue
            best_type = best["feature_type"]
            best_score = int(best.get("score") or "0")
            if score > best_score:
                best = candidate
                continue
            if feature_type == "CDS" and best_type != "CDS":
                best = candidate
                continue
            try:
                cand_len = int(end) - int(start) + 1
                best_len = int(best["end"]) - int(best["start"]) + 1
            except ValueError:
                cand_len = 0
                best_len = 0
            if cand_len > best_len:
                best = candidate
    return best


def reverse_complement_if_needed(seq: str, strand: str) -> str:
    if strand == "-":
        trans = str.maketrans("ACGTRYMKBDHVacgtrymkbdhv", "TGCAYRKMVHDBtgcayrkmvhdb")
        return seq.translate(trans)[::-1]
    return seq


def load_reference_rows() -> list[dict[str, str]]:
    workbook = load_workbook(TYPING_XLSX, read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        accession = str(values[4] or "").strip() if values else ""
        if not accession:
            continue
        accession_root = accession.split(".", 1)[0]
        rows.append(
            {
                "genus": str(values[0] or "").strip(),
                "species": str(values[1] or "").strip(),
                "virus_name": str(values[2] or "").strip(),
                "isolate": str(values[3] or "").strip(),
                "accession": accession_root,
                "available_sequence": str(values[5] or "").strip(),
                "abbrev": str(values[6] or "").strip(),
                "fasta_path": str(REF_DIR / f"{accession_root}.fasta"),
                "gff_path": str(GFF_DIR / f"{accession_root}.gff3"),
            }
        )
    return rows


def load_single_record(fasta_path: Path) -> dict[str, str]:
    header = ""
    seq_lines: list[str] = []
    with fasta_path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line in handle:
            line = line.rstrip("\n")
            if line.startswith(">"):
                if header:
                    break
                header = line[1:].strip()
                continue
            if header:
                seq_lines.append(line.strip())
    if not header:
        raise RuntimeError(f"No FASTA record found in {fasta_path}")
    return {"description": header, "sequence": "".join(seq_lines)}


def wrap_sequence(sequence: str, width: int = 80) -> str:
    return "\n".join(sequence[index:index + width] for index in range(0, len(sequence), width))


def extract_one(row: dict[str, str]) -> dict[str, str] | None:
    accession = str(row.get("accession") or "").strip()
    fasta_path = Path(str(row.get("fasta_path") or "").strip())
    gff_path = Path(str(row.get("gff_path") or "").strip())
    if not accession or not fasta_path.is_file() or not gff_path.is_file():
        return None

    feature = find_orf2_capsid_feature(gff_path)
    if feature is None:
        return None

    record = load_single_record(fasta_path)
    start = int(feature["start"])
    end = int(feature["end"])
    strand = str(feature["strand"] or "+").strip() or "+"
    subseq = record["sequence"][start - 1:end]
    subseq = reverse_complement_if_needed(subseq, strand)
    out_path = OUT_DIR / f"{accession}.orf2.fasta"
    out_path.write_text(
        f">{accession} {record['description']} | capsid ORF2 {start}-{end} ({strand})\n{wrap_sequence(subseq)}\n",
        encoding="utf-8",
    )

    result = dict(row)
    result.update(
        {
            "header": record["description"],
            "sequence_length": str(len(record["sequence"])),
            "orf2_start": str(start),
            "orf2_end": str(end),
            "orf2_strand": strand,
            "orf2_length": str(len(subseq)),
            "feature_type": feature["feature_type"],
            "product": feature.get("product", ""),
            "gene": feature.get("gene", ""),
            "note": feature.get("note", ""),
            "partial": feature.get("partial", ""),
            "orf2_fasta_path": str(out_path),
        }
    )
    return result


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest_rows = load_reference_rows()
    extracted_rows: list[dict[str, str]] = []
    missing_rows: list[dict[str, str]] = []
    merged_entries: list[str] = []

    for row in manifest_rows:
        accession = str(row.get("accession") or "").strip()
        fasta_path = Path(str(row.get("fasta_path") or "").strip())
        gff_path = Path(str(row.get("gff_path") or "").strip())
        if not accession:
            continue
        if not fasta_path.is_file():
            missing_rows.append({**row, "reason": "missing_fasta"})
            continue
        if not gff_path.is_file():
            missing_rows.append({**row, "reason": "missing_gff3"})
            continue
        result = extract_one(row)
        if result is None:
            missing_rows.append({**row, "reason": "no_orf2_capsid_feature_in_gff3"})
            continue
        extracted_rows.append(result)
        merged_entries.append(Path(result["orf2_fasta_path"]).read_text(encoding="utf-8", errors="ignore").rstrip())

    extracted_rows.sort(key=lambda item: item["accession"])
    missing_rows.sort(key=lambda item: str(item.get("accession") or ""))
    MERGED_FASTA.write_text(("\n".join(merged_entries) + "\n") if merged_entries else "", encoding="utf-8")

    fieldnames = [
        "genus",
        "species",
        "virus_name",
        "isolate",
        "accession",
        "available_sequence",
        "abbrev",
        "header",
        "sequence_length",
        "fasta_path",
        "gff_path",
        "orf2_start",
        "orf2_end",
        "orf2_strand",
        "orf2_length",
        "feature_type",
        "product",
        "gene",
        "note",
        "partial",
        "orf2_fasta_path",
    ]
    with MANIFEST_OUT.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in extracted_rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})

    with MISSING_MANIFEST.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "genus",
                "species",
                "virus_name",
                "isolate",
                "accession",
                "available_sequence",
                "abbrev",
                "header",
                "sequence_length",
                "fasta_path",
                "gff_path",
                "reason",
            ],
            delimiter="\t",
        )
        writer.writeheader()
        for row in missing_rows:
            writer.writerow({key: row.get(key, "") for key in writer.fieldnames})

    print(
        f"Extracted capsid ORF2 from {len(extracted_rows)} astroviridae references into {OUT_DIR}; "
        f"{len(missing_rows)} references were skipped"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
