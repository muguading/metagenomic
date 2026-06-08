#!/usr/bin/env python3
from __future__ import annotations

import csv
import re
from pathlib import Path
from urllib.parse import unquote

from Bio import SeqIO
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord
from openpyxl import load_workbook


ROOT = Path("/Users/wuhhh/Desktop/徐老师/代码/metagenomic")
REF_DIR = ROOT / "database" / "virus" / "enterovirus" / "reference_genomes"
GFF_DIR = REF_DIR / "gff3"
VADR_GFF_DIR = REF_DIR / "gff3_vadr"
OUT_DIR = REF_DIR / "vp1"
MERGED_FASTA = REF_DIR / "enterovirus_vp1_references.fasta"
MANIFEST = REF_DIR / "enterovirus_vp1_references.tsv"
MISSING_MANIFEST = REF_DIR / "enterovirus_vp1_missing.tsv"
TYPING_XLSX = ROOT / "database" / "virus" / "enterovirus" / "typing.xlsx"
SUPPLEMENT_MANIFEST = REF_DIR / "abcd_vp1" / "enterovirus_abcd_complete_genomes_expanded_manifest.tsv"


def parse_attributes(text: str) -> dict[str, str]:
    attrs: dict[str, str] = {}
    for item in str(text or "").strip().split(";"):
        if not item:
            continue
        if "=" in item:
            key, value = item.split("=", 1)
            attrs[key] = value
    return attrs


def load_typing_metadata() -> dict[str, dict[str, str]]:
    wb = load_workbook(TYPING_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(name or "").strip() for name in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    meta: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        accession = str(row[idx["Accession"]] or "").strip()
        if not accession or accession.startswith("partial:"):
            continue
        accession_root = accession.split(".", 1)[0]
        meta[accession_root] = {
            "accession_full": accession,
            "abbrev": str(row[idx["Abbrev."]] or "").strip(),
            "virus_name": str(row[idx["Virus Name"]] or "").strip(),
            "species": str(row[idx["Species"]] or "").strip(),
            "genus": str(row[idx["Genus"]] or "").strip(),
            "isolate": str(row[idx["Isolate"]] or "").strip(),
            "available": str(row[idx["Available"]] or "").strip(),
        }
    return meta


def load_supplement_metadata() -> dict[str, dict[str, str]]:
    if not SUPPLEMENT_MANIFEST.is_file():
        return {}
    meta: dict[str, dict[str, str]] = {}
    with SUPPLEMENT_MANIFEST.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        for row in csv.DictReader(handle, delimiter="\t"):
            accession = str(row.get("accession") or "").strip()
            if not accession:
                continue
            meta[accession] = {
                "accession_full": str(row.get("accession_full") or accession).strip(),
                "abbrev": str(row.get("abbrev") or "").strip(),
                "virus_name": str(row.get("virus_name") or "").strip(),
                "species": str(row.get("species") or "").strip(),
                "genus": str(row.get("genus") or "").strip(),
                "isolate": str(row.get("isolate") or row.get("title") or "").strip(),
                "available": str(row.get("available") or "Complete genome").strip(),
            }
    return meta


def reverse_complement_if_needed(seq: str, strand: str) -> str:
    if strand == "-":
        return str(Seq(seq).reverse_complement())
    return seq


def normalize_label(text: str) -> str:
    text = unquote(str(text or "")).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def is_vp1_annotation(product: str, gene: str, note: str) -> bool:
    raw_values = [str(product or "").strip().lower(), str(gene or "").strip().lower(), str(note or "").strip().lower()]
    normalized = [normalize_label(value) for value in raw_values]

    if any("vp1" in value for value in raw_values):
        return True
    if any(value in {"vp1", "vp1protein", "vp1polypeptide"} for value in normalized):
        return True
    if any(
        value in {
            "1d",
            "1dvp1",
            "capsidprotein1d",
            "capsidprotein1dvp1",
            "1dproteinvp1",
            "1dprotein",
        }
        for value in normalized
    ):
        return True
    if any(value.startswith("capsidprotein1d") for value in normalized):
        return True
    return False


def find_vp1_feature(gff_path: Path) -> dict[str, str] | None:
    best: dict[str, str] | None = None
    with gff_path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line in handle:
            if not line or line.startswith("#"):
                continue
            parts = line.rstrip("\n").split("\t")
            if len(parts) != 9:
                continue
            seqid, source, feature_type, start, end, score, strand, phase, attributes = parts
            attrs = parse_attributes(attributes)
            product = str(attrs.get("product") or "").strip().lower()
            gene = str(attrs.get("gene") or "").strip().lower()
            note = str(attrs.get("Note") or attrs.get("note") or "").strip().lower()
            is_vp1 = is_vp1_annotation(product, gene, note)
            if not is_vp1:
                continue
            candidate = {
                "seqid": seqid,
                "feature_type": feature_type,
                "start": start,
                "end": end,
                "strand": strand,
                "product": attrs.get("product", ""),
                "gene": attrs.get("gene", ""),
            }
            if best is None:
                best = candidate
                continue
            best_type = best["feature_type"]
            if feature_type == "mature_protein_region_of_CDS" and best_type != "mature_protein_region_of_CDS":
                best = candidate
                continue
            try:
                cand_len = int(end) - int(start) + 1
                best_len = int(best["end"]) - int(best["start"]) + 1
            except ValueError:
                cand_len = 0
                best_len = 0
            if cand_len > best_len:
                best = candidate
    return best


def load_single_record(fasta_path: Path) -> SeqRecord:
    records = list(SeqIO.parse(str(fasta_path), "fasta"))
    if not records:
        raise RuntimeError(f"No FASTA record found in {fasta_path}")
    return records[0]


def choose_gff_path(accession: str) -> Path | None:
    primary = GFF_DIR / f"{accession}.gff3"
    fallback = VADR_GFF_DIR / f"{accession}.gff3"
    if primary.is_file() and find_vp1_feature(primary) is not None:
        return primary
    if fallback.is_file() and find_vp1_feature(fallback) is not None:
        return fallback
    if primary.is_file():
        return primary
    if fallback.is_file():
        return fallback
    return None


def extract_one(fasta_path: Path, gff_path: Path) -> dict[str, str] | None:
    feature = find_vp1_feature(gff_path)
    if feature is None:
        return None
    record = load_single_record(fasta_path)
    start = int(feature["start"])
    end = int(feature["end"])
    strand = str(feature["strand"] or "+").strip() or "+"
    subseq = str(record.seq)[start - 1:end]
    subseq = reverse_complement_if_needed(subseq, strand)
    accession = fasta_path.stem
    out_record = SeqRecord(
        Seq(subseq),
        id=accession,
        description=f"{record.description} | VP1 {start}-{end} ({strand})",
    )
    out_path = OUT_DIR / f"{accession}.vp1.fasta"
    SeqIO.write([out_record], str(out_path), "fasta")
    return {
        "accession": accession,
        "header": record.description,
        "vp1_start": str(start),
        "vp1_end": str(end),
        "vp1_strand": strand,
        "vp1_length": str(len(subseq)),
        "feature_type": feature["feature_type"],
        "product": feature.get("product", ""),
        "gene": feature.get("gene", ""),
        "fasta_path": str(fasta_path),
        "gff_path": str(gff_path),
        "vp1_fasta_path": str(out_path),
    }


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    typing_meta = load_typing_metadata()
    typing_meta.update(load_supplement_metadata())
    rows: list[dict[str, str]] = []
    merged_records: list[SeqRecord] = []
    missing_rows: list[dict[str, str]] = []

    for fasta_path in sorted(REF_DIR.glob("*.fasta")):
        if fasta_path.name in {MERGED_FASTA.name, "enterovirus_typing_complete_genomes.fasta"}:
            continue
        if fasta_path.name.startswith("partial:"):
            continue
        accession = fasta_path.stem
        gff_path = choose_gff_path(accession)
        if gff_path is None:
            meta = typing_meta.get(accession, {})
            missing_rows.append(
                {
                    "accession": accession,
                    "accession_full": meta.get("accession_full", ""),
                    "abbrev": meta.get("abbrev", ""),
                    "virus_name": meta.get("virus_name", ""),
                    "species": meta.get("species", ""),
                    "genus": meta.get("genus", ""),
                    "isolate": meta.get("isolate", ""),
                    "available": meta.get("available", ""),
                    "reason": "missing_gff3",
                    "fasta_path": str(fasta_path),
                    "gff_path": str(gff_path),
                }
            )
            continue
        result = extract_one(fasta_path, gff_path)
        if result is None:
            meta = typing_meta.get(accession, {})
            missing_rows.append(
                {
                    "accession": accession,
                    "accession_full": meta.get("accession_full", ""),
                    "abbrev": meta.get("abbrev", ""),
                    "virus_name": meta.get("virus_name", ""),
                    "species": meta.get("species", ""),
                    "genus": meta.get("genus", ""),
                    "isolate": meta.get("isolate", ""),
                    "available": meta.get("available", ""),
                    "reason": "no_vp1_feature_in_gff3",
                    "fasta_path": str(fasta_path),
                    "gff_path": str(gff_path),
                }
            )
            continue
        result.update(typing_meta.get(accession, {}))
        rows.append(result)
        merged_records.extend(list(SeqIO.parse(result["vp1_fasta_path"], "fasta")))

    SeqIO.write(merged_records, str(MERGED_FASTA), "fasta")
    fieldnames = [
        "accession",
        "accession_full",
        "abbrev",
        "virus_name",
        "species",
        "genus",
        "isolate",
        "available",
        "header",
        "vp1_start",
        "vp1_end",
        "vp1_strand",
        "vp1_length",
        "feature_type",
        "product",
        "gene",
        "fasta_path",
        "gff_path",
        "vp1_fasta_path",
    ]
    with MANIFEST.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    with MISSING_MANIFEST.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "accession",
                "accession_full",
                "abbrev",
                "virus_name",
                "species",
                "genus",
                "isolate",
                "available",
                "reason",
                "fasta_path",
                "gff_path",
            ],
            delimiter="\t",
        )
        writer.writeheader()
        for row in missing_rows:
            writer.writerow(row)

    print(
        f"Extracted VP1 from {len(rows)} enterovirus references into {OUT_DIR}; "
        f"{len(missing_rows)} references were skipped"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
